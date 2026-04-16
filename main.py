#!/usr/bin/env python3
"""
Basecamp–Excel Karşılaştırma Raporu v5
───────────────────────────────────────
YENİ (v5):
  • Sabah 09:00 digest maili — API çağrısı yapmadan mevcut state özeti
  • Deadline takibi — due_on bugün/yarın olan işler sabah 09:01'de uyarı
  • /deadlines    — deadline kontrolü manuel tetikleme
  • /export.csv   — mevcut state'i CSV olarak indir
  • Dashboard — JS arama kutusu (item adına göre anlık filtre)
  • Dashboard — dark mode toggle (localStorage kalıcı)
  • Dashboard — CSV export butonu
  • Webhook log   — 30 günlük rotasyon (önceki: sadece son 10 event)

ÖNCEKİ SÜRÜMLERDEN (v4):
  Hero metrics, marka karşılaştırma, CSS trend grafiği, BC URL linkleri,
  ortalama onay süresi, webhook log, 5dk otomatik yenileme, /status sayfası

ÖNCEKİ SÜRÜMLERDEN (v1-v3):
  State kalıcılığı (STATE_DIR), hata uyarısı, /history, brand filtresi,
  dashboard önizleme, webhook güvenliği (WEBHOOK_SECRET), debounce,
  haftalık özet, süre takibi (first_seen), URL eksik kategorisi
"""

import os
import io
import json
import threading
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime, timedelta
from collections import defaultdict

import requests as req_lib

from flask import Flask, request, jsonify
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

# ─── Ortam Değişkenleri ────────────────────────────────────────────────────
BASECAMP_CLIENT_ID     = os.environ["BASECAMP_CLIENT_ID"]
BASECAMP_CLIENT_SECRET = os.environ["BASECAMP_CLIENT_SECRET"]
BASECAMP_REFRESH_TOKEN = os.environ["BASECAMP_REFRESH_TOKEN"]
BASECAMP_ACCOUNT_IDS   = ["4181631", "6168221"]

EXCEL_URL       = os.environ["EXCEL_URL"]
BREVO_API_KEY   = os.environ.get("BREVO_API_KEY", "")
RECIPIENT_EMAIL = os.environ.get("RECIPIENT_EMAIL", "ertugozerr@gmail.com")

# ─── State kalıcılığı: Railway Volume → /data, STATE_DIR=/data ─────────────
STATE_DIR        = os.environ.get("STATE_DIR", "/tmp")
STATE_FILE       = os.path.join(STATE_DIR, "bc_state.json")
WEBHOOK_LOG_FILE = os.path.join(STATE_DIR, "bc_webhook_log.json")
NOTES_FILE       = os.path.join(STATE_DIR, "bc_notes.json")
SPRINTS_FILE     = os.path.join(STATE_DIR, "bc_sprints.json")
CHANGELOG_FILE   = os.path.join(STATE_DIR, "bc_changelog.md")

# ─── Webhook güvenliği ─────────────────────────────────────────────────────
WEBHOOK_SECRET = os.environ.get("WEBHOOK_SECRET", "")

# ─── Hedef Projeler ────────────────────────────────────────────────────────
TARGET_PROJECTS = {
    "metro - dijital": "Metro",
    "hopi - sosyal medya": "Hopi",
}

PRODUKSIYON_LIST_KEYWORDS = ["prodüksiyon", "produksiyon", "production"]
SM_PM_LIST_KEYWORDS       = ["sm & pm", "sm&pm", "sm ve pm"]

# ─── Debounce ─────────────────────────────────────────────────────────────
DEBOUNCE_SECONDS = 900
_debounce_timer  = None
_debounce_lock   = threading.Lock()

# ─── Rapor kilidi ─────────────────────────────────────────────────────────
_report_lock = threading.Lock()

app = Flask(__name__)
os.makedirs(STATE_DIR, exist_ok=True)


# ══════════════════════════════════════════════════════════════════════════
#  BASECAMP API
# ══════════════════════════════════════════════════════════════════════════

def get_access_token() -> str:
    data = urllib.parse.urlencode({
        "type":          "refresh",
        "client_id":     BASECAMP_CLIENT_ID,
        "client_secret": BASECAMP_CLIENT_SECRET,
        "refresh_token": BASECAMP_REFRESH_TOKEN,
    }).encode()
    req = urllib.request.Request(
        "https://launchpad.37signals.com/authorization/token",
        data=data, method="POST",
        headers={"User-Agent": "IsOzetRaporu (ertugozerr@gmail.com)"},
    )
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read())["access_token"]


def bc_get(token: str, account_id: str, path: str) -> list:
    headers = {
        "Authorization": f"Bearer {token}",
        "User-Agent":    "IsOzetRaporu (ertugozerr@gmail.com)",
    }
    url = f"https://3.basecampapi.com/{account_id}/{path}"
    results = []
    while url:
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=30) as r:
            data = json.loads(r.read())
            results.extend(data if isinstance(data, list) else [data])
            link_header = r.headers.get("Link", "")
            url = None
            for part in link_header.split(","):
                if 'rel="next"' in part:
                    url = part.split(";")[0].strip().strip("<>")
    return results


def get_todo_info(token: str, account_id: str, bucket_id, todo_id) -> dict | None:
    try:
        items = bc_get(token, account_id, f"buckets/{bucket_id}/todos/{todo_id}.json")
        if items:
            todo = items[0]
            ln = get_todolist_name(todo)
            return {
                "completed":   bool(todo.get("completed", False)),
                "list_name":   ln,
                "produksiyon": any(k in ln for k in PRODUKSIYON_LIST_KEYWORDS),
            }
        return None
    except Exception as e:
        if "404" in str(e):
            return None
        print(f"⚠️  get_todo_info({account_id}/{bucket_id}/{todo_id}): {e}")
        return None


def get_todo_title(todo: dict) -> str:
    return (todo.get("title") or todo.get("content") or todo.get("summary") or "").strip()


def get_todolist_name(todo: dict) -> str:
    return (todo.get("parent") or {}).get("title", "").lower().strip()


def is_produksiyon(todo: dict) -> bool:
    return any(k in get_todolist_name(todo) for k in PRODUKSIYON_LIST_KEYWORDS)


def is_sm_pm(todo: dict) -> bool:
    return any(k in get_todolist_name(todo) for k in SM_PM_LIST_KEYWORDS)


def is_marka_onayinda(todo: dict) -> bool:
    return "marka onay" in get_todolist_name(todo)


# ══════════════════════════════════════════════════════════════════════════
#  EXCEL OKUMA
# ══════════════════════════════════════════════════════════════════════════

def _is_green_cell(cell) -> bool:
    try:
        fill = cell.fill
        if not fill or fill.fill_type != "solid":
            return False
        fg = fill.fgColor
        if fg.type != "rgb":
            return False
        rgb = fg.rgb
        if len(rgb) != 8 or rgb in ("00000000", "FF000000"):
            return False
        r = int(rgb[2:4], 16)
        g = int(rgb[4:6], 16)
        b = int(rgb[6:8], 16)
        return g > 150 and g > r * 1.2 and g > b
    except Exception:
        return False


def _parse_excel_bytes(content: bytes) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    tasks, seen = [], set()
    for row in ws.iter_rows(min_row=8):
        if len(row) < 4:
            continue
        cell_brand = row[2]
        cell_name  = row[3]
        cell_url   = row[4] if len(row) > 4 else None
        task_name = cell_name.value
        brand_raw = cell_brand.value
        bc_url    = str(cell_url.value) if cell_url and cell_url.value else ""
        if not task_name:
            continue
        brand = str(brand_raw).strip() if brand_raw else ""
        if brand.lower().strip() not in ("hopi", "metro"):
            continue
        cell_color = "green" if (_is_green_cell(cell_name) or _is_green_cell(cell_brand)) else "none"
        todo_id = bucket_id = url_account_id = None
        if "/todos/" in bc_url:
            raw_id = bc_url.split("/todos/")[-1].split("#")[0].strip()
            if raw_id.isdigit():
                todo_id = int(raw_id)
            try:
                parts = bc_url.split("/")
                if "buckets" in parts:
                    bi = parts.index("buckets")
                    bucket_id      = int(parts[bi + 1]) if parts[bi + 1].isdigit() else None
                    url_account_id = parts[3] if parts[3].isdigit() else None
            except Exception:
                pass
        key = todo_id if todo_id else str(task_name).strip().lower()
        if key in seen:
            continue
        seen.add(key)
        tasks.append({
            "name": str(task_name).strip(), "brand": brand,
            "todo_id": todo_id, "bucket_id": bucket_id,
            "url_account_id": url_account_id, "cell_color": cell_color,
        })
    return tasks


def read_excel_tasks() -> list[dict]:
    sep = "&" if "?" in EXCEL_URL else "?"
    download_url = EXCEL_URL + sep + "download=1"
    session = req_lib.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})
    session.get(EXCEL_URL, timeout=30, allow_redirects=True)
    r = session.get(download_url, timeout=30, allow_redirects=True)
    content = r.content
    if content[:2] != b"PK":
        raise ValueError(f"xlsx değil: {content[:120].decode('utf-8', errors='replace')}")
    return _parse_excel_bytes(content)


# ══════════════════════════════════════════════════════════════════════════
#  DURUM DOSYASI
# ══════════════════════════════════════════════════════════════════════════

HISTORY_KEEP_DAYS = 14


def _item_names(lst: list) -> set:
    return {i["name"] if isinstance(i, dict) else i for i in lst}


def _as_items(lst: list) -> list[dict]:
    return [i if isinstance(i, dict) else {"name": str(i), "brand": ""} for i in lst]


def _to_state_items(items: list) -> list[dict]:
    """Item dict'lerini state formatına çevirir. BC URL bilgisini de saklar."""
    result = []
    for t in items:
        result.append({
            "name":           t["name"],
            "brand":          t.get("brand", ""),
            "todo_id":        t.get("todo_id") or t.get("id"),
            "bucket_id":      t.get("bucket_id"),
            "url_account_id": t.get("url_account_id"),
        })
    return result


def _bc_url(item: dict) -> str:
    """State item'ından Basecamp URL'si oluşturur."""
    tid  = item.get("todo_id")
    bid  = item.get("bucket_id")
    acct = item.get("url_account_id")
    if tid and bid and acct:
        return f"https://3.basecamp.com/{acct}/buckets/{bid}/todos/{tid}"
    return ""


def load_state() -> dict:
    try:
        with open(STATE_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_state(sil, yesile, renksiz, ekle, url_eksik, timestamp) -> dict:
    today_str = datetime.now().strftime("%Y-%m-%d")
    existing  = load_state()
    old_fs    = existing.get("first_seen", {})
    history   = list(existing.get("history", []))

    new_fs = {}
    for cat, items in [("sil", sil), ("yesile", yesile), ("renksiz", renksiz), ("ekle", ekle)]:
        for item in items:
            key = f"{cat}:{item['name']}"
            new_fs[key] = old_fs.get(key, today_str)

    prev_yesile = _item_names(existing.get("yesile", []))
    prev_sil    = _item_names(existing.get("sil", []))
    curr_yesile = {t["name"] for t in yesile}
    curr_sil    = {t["name"] for t in sil}

    history.append({
        "date":          today_str,
        "time":          timestamp,
        "sil_count":     len(sil),
        "yesile_count":  len(yesile),
        "renksiz_count": len(renksiz),
        "ekle_count":    len(ekle),
        "completed":     list(curr_sil - prev_sil),
        "new_onay":      list(curr_yesile - prev_yesile),
    })
    cutoff  = (datetime.now() - timedelta(days=HISTORY_KEEP_DAYS)).strftime("%Y-%m-%d")
    history = [h for h in history if h.get("date", "") >= cutoff]

    state = {
        "timestamp":  timestamp,
        "sil":        _to_state_items(sil),
        "yesile":     _to_state_items(yesile),
        "renksiz":    _to_state_items(renksiz),
        "ekle":       _to_state_items(ekle),
        "url_eksik":  _to_state_items(url_eksik),
        "first_seen": new_fs,
        "history":    history,
        "last_error": None,
    }
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️  State kayıt: {e}")
    return new_fs


def set_last_error(error: str):
    try:
        state = load_state()
        state["last_error"] = {"message": error, "time": datetime.now().strftime("%d.%m.%Y %H:%M")}
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def get_days_in_category(first_seen: dict, category: str, name: str) -> int | None:
    date_str = first_seen.get(f"{category}:{name}")
    if not date_str:
        return None
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
        return (datetime.now().date() - d).days
    except Exception:
        return None


def compute_changes(prev, sil, yesile, renksiz, ekle) -> list[str]:
    if not prev:
        return []
    changes = []
    curr_sil     = {t["name"] for t in sil}
    curr_yesile  = {t["name"] for t in yesile}
    curr_renksiz = {t["name"] for t in renksiz}
    curr_ekle    = {t["name"] for t in ekle}
    prev_sil     = _item_names(prev.get("sil", []))
    prev_yesile  = _item_names(prev.get("yesile", []))
    prev_renksiz = _item_names(prev.get("renksiz", []))
    prev_ekle    = _item_names(prev.get("ekle", []))

    if yeni := curr_sil - prev_sil:
        changes.append("Yeni tamamlandı: " + ", ".join(sorted(yeni)))
    if silindi := prev_sil - curr_sil:
        changes.append("Excel'den silindi: " + ", ".join(sorted(silindi)))
    if onay := curr_yesile - prev_yesile:
        changes.append("Marka onayına geldi: " + ", ".join(sorted(onay)))
    if cikti := curr_renksiz - prev_renksiz:
        changes.append("Onaydan çıktı (renksiz yap): " + ", ".join(sorted(cikti)))
    if ekle_new := curr_ekle - prev_ekle:
        changes.append("Basecamp'te yeni iş: " + ", ".join(sorted(ekle_new)))
    return changes


# ══════════════════════════════════════════════════════════════════════════
#  WEBHOOK LOG
# ══════════════════════════════════════════════════════════════════════════

WEBHOOK_LOG_KEEP_DAYS = 30
WEBHOOK_LOG_MAX       = 2000   # maksimum satır (dosya şişmesin)


def log_webhook_event(kind: str, status: str):
    """Webhook event'lerini 30 gün saklayan döngüsel log."""
    try:
        try:
            with open(WEBHOOK_LOG_FILE, encoding="utf-8") as f:
                log = json.load(f)
        except Exception:
            log = []
        log.insert(0, {
            "time":   datetime.now().strftime("%d.%m.%Y %H:%M"),
            "kind":   kind,
            "status": status,
        })
        # 30 günden eski entryleri at
        cutoff = datetime.now() - timedelta(days=WEBHOOK_LOG_KEEP_DAYS)
        def _parse(t):
            try:
                return datetime.strptime(t, "%d.%m.%Y %H:%M")
            except Exception:
                return datetime.min
        log = [e for e in log if _parse(e.get("time", "")) >= cutoff]
        log = log[:WEBHOOK_LOG_MAX]   # güvenlik tavanı
        with open(WEBHOOK_LOG_FILE, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️  Webhook log: {e}")


def load_webhook_log() -> list:
    try:
        with open(WEBHOOK_LOG_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


# ══════════════════════════════════════════════════════════════════════════
#  NOTLAR
# ══════════════════════════════════════════════════════════════════════════

def load_notes() -> dict:
    """{'İş Adı': {'text': '...', 'time': 'gg.aa.yyyy HH:MM'}}"""
    try:
        with open(NOTES_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_note(item_name: str, note_text: str):
    notes = load_notes()
    if note_text.strip():
        notes[item_name] = {
            "text": note_text.strip(),
            "time": datetime.now().strftime("%d.%m.%Y %H:%M"),
        }
    else:
        notes.pop(item_name, None)   # boş not = sil
    try:
        with open(NOTES_FILE, "w", encoding="utf-8") as f:
            json.dump(notes, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️  Not kayıt: {e}")


# ══════════════════════════════════════════════════════════════════════════
#  SPRINT / KAMPANYA
# ══════════════════════════════════════════════════════════════════════════

def load_sprints() -> list:
    try:
        with open(SPRINTS_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def save_sprint(name: str, start: str, end: str):
    sprints = load_sprints()
    sprints.append({
        "name":    name,
        "start":   start,   # YYYY-MM-DD
        "end":     end,
        "created": datetime.now().strftime("%d.%m.%Y %H:%M"),
    })
    try:
        with open(SPRINTS_FILE, "w", encoding="utf-8") as f:
            json.dump(sprints, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️  Sprint kayıt: {e}")


def get_active_sprint() -> dict | None:
    today = datetime.now().strftime("%Y-%m-%d")
    for s in reversed(load_sprints()):
        if s.get("start", "") <= today <= s.get("end", ""):
            return s
    return None


def build_sprint_page(sprint: dict, history: list) -> str:
    """Bir sprint / kampanya için özet HTML sayfası."""
    start = sprint.get("start", "")
    end   = sprint.get("end", "")
    name  = sprint.get("name", "Sprint")

    entries = [h for h in history if start <= h.get("date", "") <= end]
    entries_sorted = sorted(entries, key=lambda x: x.get("date", ""))

    total_completed = len({n for h in entries for n in h.get("completed", [])})
    total_new_onay  = len({n for h in entries for n in h.get("new_onay", [])})
    total_ekle      = sum(h.get("ekle_count", 0) for h in entries)
    report_count    = len(entries)

    day_rows = ""
    for h in entries_sorted:
        compl = len(h.get("completed", []))
        onay  = len(h.get("new_onay", []))
        day_rows += (
            f"<tr>"
            f"<td style='padding:7px 12px;border-bottom:1px solid #eee;font-size:13px;'>{h.get('date','')}</td>"
            f"<td style='padding:7px 12px;border-bottom:1px solid #eee;text-align:center;'>"
            f"<span style='background:#00b050;color:#fff;border-radius:3px;padding:1px 8px;font-size:12px;'>{h.get('yesile_count',0)}</span></td>"
            f"<td style='padding:7px 12px;border-bottom:1px solid #eee;text-align:center;'>"
            f"<span style='background:#e53935;color:#fff;border-radius:3px;padding:1px 8px;font-size:12px;'>{h.get('sil_count',0)}</span></td>"
            f"<td style='padding:7px 12px;border-bottom:1px solid #eee;text-align:center;font-weight:bold;color:#00b050;'>"
            f"{'+ ' + str(compl) if compl else '—'}</td>"
            f"<td style='padding:7px 12px;border-bottom:1px solid #eee;text-align:center;font-weight:bold;color:#1976d2;'>"
            f"{'+ ' + str(onay) if onay else '—'}</td>"
            f"</tr>"
        )
    if not day_rows:
        day_rows = "<tr><td colspan='5' style='padding:16px;text-align:center;color:#888;'>Bu tarih aralığında rapor yok.</td></tr>"

    all_sprints = load_sprints()
    sprint_links = ""
    for i, s in enumerate(reversed(all_sprints)):
        s_start = s.get("start", "")
        s_end   = s.get("end", "")
        s_name  = s.get("name", "Sprint")
        active  = s_start == start and s_end == end
        bg = "#1976d2" if active else "#e0e0e0"
        fg = "#fff" if active else "#333"
        sprint_links += (
            f"<a href='/sprint?start={s_start}&end={s_end}' "
            f"style='background:{bg};color:{fg};padding:5px 12px;border-radius:5px;"
            f"font-size:12px;text-decoration:none;margin-right:6px;'>{s_name}</a>"
        )

    return f"""<!DOCTYPE html>
<html><head>
  <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Sprint · {name}</title>
  <style>body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif;background:#f0f2f5;margin:0;padding:20px;}}
  .wrap{{max-width:820px;margin:0 auto;}} a{{color:#1976d2;text-decoration:none;}} a:hover{{text-decoration:underline;}}</style>
</head>
<body><div class="wrap">
  <div style='background:linear-gradient(135deg,#1a1a2e,#16213e);color:#fff;border-radius:12px;padding:22px 26px;margin-bottom:20px;'>
    <div style='font-size:11px;text-transform:uppercase;letter-spacing:1px;opacity:.7;'>PunchBBDO — Sprint Takibi</div>
    <div style='font-size:24px;font-weight:bold;margin-top:4px;'>🏃 {name}</div>
    <div style='font-size:13px;opacity:.7;margin-top:4px;'>{start} → {end} &nbsp;·&nbsp; {report_count} rapor</div>
    <div style='margin-top:12px;'>{sprint_links}</div>
  </div>
  <div style='display:flex;gap:10px;margin-bottom:20px;flex-wrap:wrap;'>
    <div style='flex:1;min-width:110px;background:#fff;border-radius:8px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.08);text-align:center;border-top:3px solid #00b050;'>
      <div style='font-size:30px;font-weight:bold;color:#00b050;'>{total_completed}</div>
      <div style='font-size:12px;color:#666;margin-top:4px;'>Tamamlandı</div></div>
    <div style='flex:1;min-width:110px;background:#fff;border-radius:8px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.08);text-align:center;border-top:3px solid #1976d2;'>
      <div style='font-size:30px;font-weight:bold;color:#1976d2;'>{total_new_onay}</div>
      <div style='font-size:12px;color:#666;margin-top:4px;'>Marka Onayına Geldi</div></div>
    <div style='flex:1;min-width:110px;background:#fff;border-radius:8px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.08);text-align:center;border-top:3px solid #9c27b0;'>
      <div style='font-size:30px;font-weight:bold;color:#9c27b0;'>{total_ekle}</div>
      <div style='font-size:12px;color:#666;margin-top:4px;'>Excel'e Eklendi</div></div>
    <div style='flex:1;min-width:110px;background:#fff;border-radius:8px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.08);text-align:center;border-top:3px solid #607d8b;'>
      <div style='font-size:30px;font-weight:bold;color:#607d8b;'>{report_count}</div>
      <div style='font-size:12px;color:#666;margin-top:4px;'>Rapor Çalıştı</div></div>
  </div>
  <div style='background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.1);overflow:hidden;margin-bottom:20px;'>
    <div style='background:#1a1a2e;color:#fff;padding:12px 16px;font-weight:bold;'>📊 Günlük Detay</div>
    <table style='width:100%;border-collapse:collapse;'>
      <tr style='background:#fafafa;'>
        <th style='padding:7px 12px;text-align:left;font-size:12px;color:#888;'>Tarih</th>
        <th style='padding:7px 12px;text-align:center;font-size:12px;color:#888;'>Onayda</th>
        <th style='padding:7px 12px;text-align:center;font-size:12px;color:#888;'>Silinecek</th>
        <th style='padding:7px 12px;text-align:center;font-size:12px;color:#888;'>Tamamlandı</th>
        <th style='padding:7px 12px;text-align:center;font-size:12px;color:#888;'>Yeni Onay</th>
      </tr>{day_rows}
    </table>
  </div>
  <div style='text-align:center;font-size:11px;color:#aaa;'>
    <a href='/dashboard'>← Dashboard</a> &nbsp;·&nbsp;
    <a href='/sprint/new'>+ Yeni Sprint</a>
  </div>
</div></body></html>"""


# ══════════════════════════════════════════════════════════════════════════
#  CHANGELOG
# ══════════════════════════════════════════════════════════════════════════

def append_changelog(changes: list, today: str, trigger: str):
    """Her rapordaki değişiklikleri kalıcı markdown dosyasına ekler."""
    if not changes:
        return
    try:
        lines = [f"\n## {today}  _(tetikleyen: {trigger})_\n"]
        for c in changes:
            lines.append(f"- {c}\n")
        with open(CHANGELOG_FILE, "a", encoding="utf-8") as f:
            f.writelines(lines)
    except Exception as e:
        print(f"⚠️  Changelog: {e}")


# ══════════════════════════════════════════════════════════════════════════
#  ISI HARİTASI + TAHMİN MOTORU
# ══════════════════════════════════════════════════════════════════════════

def build_heatmap_html(webhook_log: list) -> str:
    """Webhook log → gün × saat ısı haritası HTML."""
    # 7 × 24 matris, sadece "scheduled" event'ler sayılır
    matrix = [[0] * 24 for _ in range(7)]   # [weekday][hour]
    for ev in webhook_log:
        if ev.get("status") == "ignored":
            continue
        time_str = ev.get("time", "")
        try:
            dt = datetime.strptime(time_str, "%d.%m.%Y %H:%M")
            matrix[dt.weekday()][dt.hour] += 1
        except Exception:
            pass

    max_val = max((matrix[d][h] for d in range(7) for h in range(24)), default=1)
    max_val = max(max_val, 1)

    day_labels = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]
    # Sadece iş saatlerini göster: 8–20
    show_hours = list(range(8, 21))

    header = "<tr><th style='padding:3px 6px;font-size:10px;color:#888;text-align:right;'>Saat →</th>"
    for h in show_hours:
        header += f"<th style='padding:3px 4px;font-size:10px;color:#888;text-align:center;'>{h}</th>"
    header += "</tr>"

    rows_html = ""
    for d in range(7):
        row = f"<tr><td style='padding:3px 6px;font-size:11px;color:#555;font-weight:bold;text-align:right;'>{day_labels[d]}</td>"
        for h in show_hours:
            v = matrix[d][h]
            intensity = v / max_val
            # Renk: 0=beyaz, yüksek=koyu mavi
            r = int(255 - intensity * 25)
            g = int(255 - intensity * 100)
            b = int(255 - intensity * 0)
            # Aslında mavi ton: beyazdan koyu maviye
            r2 = int(255 - intensity * 200)
            g2 = int(255 - intensity * 130)
            b2 = 255
            bg = f"rgb({r2},{g2},{b2})" if v > 0 else "#f5f5f5"
            fg = "#fff" if intensity > 0.6 else "#333"
            title = f"{day_labels[d]} {h}:00 — {v} event"
            row += (f"<td title='{title}' style='padding:4px 3px;text-align:center;"
                    f"background:{bg};color:{fg};font-size:10px;border-radius:3px;min-width:22px;'>"
                    f"{'  ' if v == 0 else str(v)}</td>")
        row += "</tr>"
        rows_html += row

    if max_val == 1 and all(matrix[d][h] == 0 for d in range(7) for h in range(24)):
        return ""  # log boşsa section gösterme

    return (
        f"<div style='background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.1);"
        f"padding:16px;margin-bottom:20px;overflow-x:auto;'>"
        f"<div style='font-weight:bold;font-size:14px;color:#333;margin-bottom:10px;'>"
        f"🌡️ Aktivite Isı Haritası <span style='font-size:11px;color:#aaa;font-weight:normal;'>"
        f"(webhook event'ler, iş saatleri 08–20)</span></div>"
        f"<table style='border-collapse:separate;border-spacing:2px;'>"
        f"{header}{rows_html}</table>"
        f"<div style='margin-top:8px;font-size:11px;color:#aaa;'>"
        f"Beyaz = 0, koyu mavi = yoğun aktivite</div></div>"
    )


def build_forecast_widget(history: list) -> str:
    """Son 4 haftanın ortalamasından bu hafta tahmini."""
    if len(history) < 3:
        return ""

    # history entry'lerine hafta numarası ekle
    weekly: dict = defaultdict(lambda: {"yesile": [], "sil": [], "ekle": []})
    for h in history:
        date_str = h.get("date", "")
        try:
            dt      = datetime.strptime(date_str, "%Y-%m-%d")
            week_key = dt.strftime("%Y-W%W")
            weekly[week_key]["yesile"].append(h.get("yesile_count", 0))
            weekly[week_key]["sil"].append(h.get("sil_count", 0))
            weekly[week_key]["ekle"].append(h.get("ekle_count", 0))
        except Exception:
            pass

    # Son 4 tam hafta (bu haftayı hariç tut)
    current_week = datetime.now().strftime("%Y-W%W")
    past_weeks = sorted([k for k in weekly if k != current_week], reverse=True)[:4]
    if not past_weeks:
        return ""

    def avg(cat):
        vals = [sum(weekly[w][cat]) / max(len(weekly[w][cat]), 1) for w in past_weeks]
        return round(sum(vals) / len(vals), 1)

    avg_yesile = avg("yesile")
    avg_sil    = avg("sil")
    avg_ekle   = avg("ekle")
    basis      = len(past_weeks)

    return (
        f"<div style='background:#fff;border-radius:10px;padding:16px 20px;"
        f"box-shadow:0 1px 4px rgba(0,0,0,.1);margin-bottom:20px;"
        f"border-left:4px solid #7b1fa2;'>"
        f"<div style='font-weight:bold;font-size:14px;color:#7b1fa2;margin-bottom:10px;'>"
        f"🔮 Bu Hafta Tahmini <span style='font-size:11px;color:#aaa;font-weight:normal;'>"
        f"(son {basis} haftanın günlük ortalaması)</span></div>"
        f"<div style='display:flex;gap:16px;flex-wrap:wrap;'>"
        f"<div style='text-align:center;'>"
        f"<div style='font-size:26px;font-weight:bold;color:#00b050;'>{avg_yesile}</div>"
        f"<div style='font-size:11px;color:#888;'>Marka Onayına/gün</div></div>"
        f"<div style='text-align:center;'>"
        f"<div style='font-size:26px;font-weight:bold;color:#e53935;'>{avg_sil}</div>"
        f"<div style='font-size:11px;color:#888;'>Tamamlanacak/gün</div></div>"
        f"<div style='text-align:center;'>"
        f"<div style='font-size:26px;font-weight:bold;color:#1976d2;'>{avg_ekle}</div>"
        f"<div style='font-size:11px;color:#888;'>Yeni İş/gün</div></div>"
        f"</div></div>"
    )


def _build_notes_list(notes: dict) -> str:
    """Mevcut notlar listesini HTML olarak döndürür (f-string dışında üretir)."""
    if not notes:
        return ""
    rows = ""
    for item_n, nd in notes.items():
        # URL-safe delete link: & → %26, ' → %27
        safe_key = urllib.parse.quote(item_n, safe="")
        rows += (
            f"<div style='display:flex;justify-content:space-between;align-items:center;"
            f"padding:6px 0;border-bottom:1px solid #f5f5f5;'>"
            f"<div><b style='font-size:13px;'>{item_n}</b>"
            f"<span style='color:#7b1fa2;font-size:12px;'> — {nd.get('text','')}</span>"
            f"<span style='color:#bbb;font-size:11px;'> ({nd.get('time','')})</span></div>"
            f"<a href='/note?delete={safe_key}' style='color:#e53935;font-size:11px;'>sil</a>"
            f"</div>"
        )
    return (
        "<div style='margin-top:14px;'>"
        "<div style='font-size:12px;color:#888;margin-bottom:6px;'>Mevcut notlar:</div>"
        + rows
        + "</div>"
    )


# ══════════════════════════════════════════════════════════════════════════
#  DASHBOARD YARDIMCILARİ
# ══════════════════════════════════════════════════════════════════════════

def next_report_countdown() -> str:
    """Sonraki haftaiçi 18:00 Türkiye saatine kalan süreyi döndürür."""
    # Türkiye = UTC+3
    now = datetime.utcnow() + timedelta(hours=3)
    target = now.replace(hour=18, minute=0, second=0, microsecond=0)

    if now >= target or now.weekday() >= 5:
        days_ahead = 1
        while True:
            candidate = now + timedelta(days=days_ahead)
            if candidate.weekday() < 5:
                target = candidate.replace(hour=18, minute=0, second=0, microsecond=0)
                break
            days_ahead += 1

    delta      = target - now
    total_secs = int(delta.total_seconds())
    hours      = total_secs // 3600
    minutes    = (total_secs % 3600) // 60
    day_names  = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]
    day        = day_names[target.weekday()]

    if delta.days >= 1:
        return f"{day} 18:00 ({hours} saat sonra)"
    elif hours > 0:
        return f"{hours} saat {minutes} dk sonra"
    else:
        return f"{minutes} dk sonra"


def avg_onay_days(first_seen: dict) -> str:
    """Şu an yesile listesindeki itemların ortalama bekleme süresini döndürür."""
    days_list = []
    for key, date_str in first_seen.items():
        if not key.startswith("yesile:"):
            continue
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d").date()
            days_list.append((datetime.now().date() - d).days)
        except Exception:
            pass
    if not days_list:
        return "—"
    avg = sum(days_list) / len(days_list)
    return f"{avg:.1f} gün"


# ══════════════════════════════════════════════════════════════════════════
#  RAPOR METNİ
# ══════════════════════════════════════════════════════════════════════════

def _days_label(days: int | None) -> str:
    if days is None or days == 0:
        return ""
    return " (1 gün)" if days == 1 else f" ({days} gündür)"


def build_report(yesile_boya, renksiz_yap, sil_listesi, ekle_listesi, url_eksik,
                 today, excel_error="", changes=None, first_seen=None) -> str:
    fs = first_seen or {}

    def fmt(items, cat):
        if not items:
            return ["  (Yok)"]
        return [
            f"  - {t['name']} — {t.get('brand','')}{_days_label(get_days_in_category(fs, cat, t['name']))}"
            + (" 🟢 (yeşile boya)" if t.get("yesile_boya") else "")
            for t in items
        ]

    lines = [f"📋 EXCEL GÜNCELLEME TALİMATLARI — {today}", ""]
    if excel_error:
        lines += [f"⚠️  Excel okunamadı: {excel_error}", ""]
    if changes:
        lines += ["🔄 SON RAPORDAN DEĞİŞİKLİKLER:"] + [f"  • {c}" for c in changes] + [""]
    lines += ["🗑️  SİL (Basecamp'te tamamlandı):"] + fmt(sil_listesi, "sil") + [""]
    lines += ["🟢 YEŞİLE BOYA (Marka Onayında — Excel'de henüz yeşil değil):"] + fmt(yesile_boya, "yesile") + [""]
    lines += ["⬜ RENKSİZ YAP (Artık Marka Onayında değil — Excel'de hâlâ yeşil):"] + fmt(renksiz_yap, "renksiz") + [""]
    lines += ["➕ EXCEL'E EKLE (Basecamp'te var, Excel'de yok):"] + fmt(ekle_listesi, "ekle")
    if url_eksik:
        lines += ["", "🔗 BASECAMP URL EKSİK:"] + fmt(url_eksik, "url_eksik")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════
#  HTML MAİL — GÜNLÜK
# ══════════════════════════════════════════════════════════════════════════

def _dur_badge_html(days: int | None) -> str:
    if days is None or days == 0:
        return ""
    color = "#9e9e9e" if days <= 2 else ("#f57c00" if days <= 6 else "#c62828")
    return (f"&nbsp;<span style='background:{color};color:#fff;padding:1px 6px;"
            f"border-radius:3px;font-size:11px;font-weight:bold;'>{days}&nbsp;gün</span>")


def _html_card(title, color, icon, items, category, first_seen):
    fs = first_seen or {}
    if not items:
        rows = "<li style='color:#888;font-style:italic;'>Yok</li>"
    else:
        rows = ""
        for t in items:
            days  = get_days_in_category(fs, category, t["name"])
            badge = _dur_badge_html(days)
            note  = (" &nbsp;<span style='background:#00b050;color:#fff;padding:1px 6px;"
                     "border-radius:3px;font-size:11px;'>yeşile boya</span>"
                     if t.get("yesile_boya") else "")
            rows += (f"<li style='padding:4px 0;border-bottom:1px solid #f0f0f0;'>"
                     f"<b>{t['name']}</b> <span style='color:#888;font-size:12px;'>"
                     f"— {t.get('brand','')}</span>{badge}{note}</li>")
    return (f"<div style='margin:12px 0;border-radius:8px;overflow:hidden;"
            f"box-shadow:0 1px 4px rgba(0,0,0,.08);border-left:5px solid {color};background:#fff;'>"
            f"<div style='background:{color};color:#fff;padding:10px 16px;"
            f"font-weight:bold;font-size:14px;'>{icon}&nbsp;&nbsp;{title}</div>"
            f"<ul style='margin:0;padding:12px 16px 12px 32px;list-style:disc;'>{rows}</ul></div>")


def build_html_report(yesile_boya, renksiz_yap, sil_listesi, ekle_listesi, url_eksik,
                      today, excel_error="", changes=None, first_seen=None) -> str:
    fs = first_seen or {}
    changes_block = ""
    if changes:
        ih = "".join(f"<li>{c}</li>" for c in changes)
        changes_block = (f"<div style='margin:12px 0;border-radius:8px;background:#fff8e1;"
                         f"border:1px solid #ffe082;padding:12px 16px;'>"
                         f"<div style='font-weight:bold;color:#f57c00;margin-bottom:6px;'>"
                         f"🔄 Son Rapordan Değişiklikler</div>"
                         f"<ul style='margin:0;padding-left:20px;color:#555;'>{ih}</ul></div>")
    error_block = (f"<div style='background:#fff3f3;border:1px solid #ffcdd2;"
                   f"border-radius:8px;padding:10px 16px;margin:12px 0;color:#c62828;'>"
                   f"⚠️ Excel okunamadı: {excel_error}</div>") if excel_error else ""
    url_block = (_html_card("BASECAMP URL EKSİK", "#9c27b0", "🔗", url_eksik, "url_eksik", fs)
                 if url_eksik else "")
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style='font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif;background:#f5f5f5;margin:0;padding:20px;'>
  <div style='max-width:600px;margin:0 auto;'>
    <div style='background:linear-gradient(135deg,#1a1a2e,#16213e);color:#fff;border-radius:10px;padding:20px 24px;margin-bottom:16px;'>
      <div style='font-size:11px;text-transform:uppercase;letter-spacing:1px;opacity:.7;margin-bottom:4px;'>PunchBBDO — Excel Takip</div>
      <div style='font-size:22px;font-weight:bold;'>📋 Güncelleme Talimatları</div>
      <div style='font-size:13px;opacity:.8;margin-top:4px;'>{today}</div>
    </div>
    {error_block}{changes_block}
    {_html_card("SİL — Basecamp'te tamamlandı","#e53935","🗑️",sil_listesi,"sil",fs)}
    {_html_card("YEŞİLE BOYA — Marka Onayında, Excel'de henüz yeşil değil","#00b050","🟢",yesile_boya,"yesile",fs)}
    {_html_card("RENKSİZ YAP — Artık Marka Onayında değil, Excel'de hâlâ yeşil","#757575","⬜",renksiz_yap,"renksiz",fs)}
    {_html_card("EXCEL'E EKLE — Basecamp'te var, Excel'de yok","#1976d2","➕",ekle_listesi,"ekle",fs)}
    {url_block}
    <div style='text-align:center;font-size:11px;color:#aaa;margin-top:20px;'>Otomatik rapor · bc-takip-production.up.railway.app</div>
  </div>
</body></html>"""


# ══════════════════════════════════════════════════════════════════════════
#  HTML MAİL — HAFTALIK ÖZET
# ══════════════════════════════════════════════════════════════════════════

def build_weekly_html(state: dict) -> str:
    today      = datetime.now().strftime("%d.%m.%Y")
    history    = state.get("history", [])
    first_seen = state.get("first_seen", {})
    cutoff     = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    week       = [h for h in history if h.get("date", "") >= cutoff]

    total_completed = list({n for h in week for n in h.get("completed", [])})
    total_new_onay  = list({n for h in week for n in h.get("new_onay", [])})

    long_waiters = []
    for key, date_str in first_seen.items():
        if not key.startswith("yesile:"):
            continue
        name = key[len("yesile:"):]
        try:
            d    = datetime.strptime(date_str, "%Y-%m-%d").date()
            days = (datetime.now().date() - d).days
            if days >= 3:
                long_waiters.append((name, days))
        except Exception:
            pass
    long_waiters.sort(key=lambda x: -x[1])

    day_rows = ""
    for h in sorted(week, key=lambda x: x.get("date", ""), reverse=True):
        compl = len(h.get("completed", []))
        day_rows += (f"<tr>"
                     f"<td style='padding:6px 10px;border-bottom:1px solid #eee;'>{h.get('date','')}</td>"
                     f"<td style='padding:6px 10px;border-bottom:1px solid #eee;text-align:center;'>{h.get('yesile_count',0)}</td>"
                     f"<td style='padding:6px 10px;border-bottom:1px solid #eee;text-align:center;'>{h.get('sil_count',0)}</td>"
                     f"<td style='padding:6px 10px;border-bottom:1px solid #eee;text-align:center;'>{h.get('ekle_count',0)}</td>"
                     f"<td style='padding:6px 10px;border-bottom:1px solid #eee;text-align:center;"
                     f"font-weight:bold;color:#00b050;'>{'+ ' + str(compl) if compl else '—'}</td></tr>")
    if not day_rows:
        day_rows = "<tr><td colspan='5' style='padding:12px;text-align:center;color:#888;'>Henüz veri yok</td></tr>"

    waiter_rows = "".join(
        f"<tr><td style='padding:6px 10px;border-bottom:1px solid #eee;'>{n}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eee;text-align:center;"
        f"color:{'#c62828' if d>=7 else '#f57c00'};font-weight:bold;'>{d} gün</td></tr>"
        for n, d in long_waiters
    )
    waiter_section = (
        f"<div style='background:#fff;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.08);"
        f"margin-bottom:16px;overflow:hidden;border-left:5px solid #f57c00;'>"
        f"<div style='background:#f57c00;color:#fff;padding:10px 16px;font-weight:bold;'>"
        f"⏰ Uzun Süredir Marka Onayında ({len(long_waiters)} iş)</div>"
        f"<table style='width:100%;border-collapse:collapse;'>"
        f"<tr style='background:#fafafa;'><th style='padding:6px 10px;text-align:left;font-size:12px;color:#888;'>İş Adı</th>"
        f"<th style='padding:6px 10px;text-align:center;font-size:12px;color:#888;'>Süre</th></tr>"
        f"{waiter_rows}</table></div>"
    ) if long_waiters else ""

    c_html = ("".join(f"<li>{n}</li>" for n in sorted(total_completed))
              or "<li style='color:#888;font-style:italic;'>Yok</li>")
    o_html = ("".join(f"<li>{n}</li>" for n in sorted(total_new_onay))
              or "<li style='color:#888;font-style:italic;'>Yok</li>")
    lw_col = "#c62828" if long_waiters else "#4caf50"

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style='font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif;background:#f5f5f5;margin:0;padding:20px;'>
  <div style='max-width:620px;margin:0 auto;'>
    <div style='background:linear-gradient(135deg,#1a1a2e,#16213e);color:#fff;border-radius:10px;padding:20px 24px;margin-bottom:16px;'>
      <div style='font-size:11px;text-transform:uppercase;letter-spacing:1px;opacity:.7;margin-bottom:4px;'>PunchBBDO — Excel Takip</div>
      <div style='font-size:22px;font-weight:bold;'>📅 Haftalık Özet</div>
      <div style='font-size:13px;opacity:.8;margin-top:4px;'>{today}</div>
    </div>
    <div style='display:flex;gap:10px;margin-bottom:16px;flex-wrap:wrap;'>
      <div style='flex:1;min-width:120px;background:#fff;border-radius:8px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.08);text-align:center;'>
        <div style='font-size:32px;font-weight:bold;color:#00b050;'>{len(total_completed)}</div>
        <div style='font-size:12px;color:#666;margin-top:4px;'>Bu Hafta Tamamlanan</div>
      </div>
      <div style='flex:1;min-width:120px;background:#fff;border-radius:8px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.08);text-align:center;'>
        <div style='font-size:32px;font-weight:bold;color:#1976d2;'>{len(total_new_onay)}</div>
        <div style='font-size:12px;color:#666;margin-top:4px;'>Marka Onayına Geldi</div>
      </div>
      <div style='flex:1;min-width:120px;background:#fff;border-radius:8px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.08);text-align:center;'>
        <div style='font-size:32px;font-weight:bold;color:{lw_col};'>{len(long_waiters)}</div>
        <div style='font-size:12px;color:#666;margin-top:4px;'>3+ Gündür Bekleyen</div>
      </div>
    </div>
    {waiter_section}
    <div style='background:#fff;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.08);margin-bottom:16px;overflow:hidden;'>
      <div style='background:#1a1a2e;color:#fff;padding:10px 16px;font-weight:bold;'>📊 Günlük İstatistikler</div>
      <table style='width:100%;border-collapse:collapse;'>
        <tr style='background:#fafafa;'>
          <th style='padding:6px 10px;text-align:left;font-size:12px;color:#888;'>Tarih</th>
          <th style='padding:6px 10px;text-align:center;font-size:12px;color:#888;'>Onayda</th>
          <th style='padding:6px 10px;text-align:center;font-size:12px;color:#888;'>Silinecek</th>
          <th style='padding:6px 10px;text-align:center;font-size:12px;color:#888;'>Eklenecek</th>
          <th style='padding:6px 10px;text-align:center;font-size:12px;color:#888;'>Tamamlandı</th>
        </tr>{day_rows}
      </table>
    </div>
    <div style='display:flex;gap:10px;flex-wrap:wrap;'>
      <div style='flex:1;min-width:200px;background:#fff;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.08);padding:14px;border-left:5px solid #00b050;'>
        <div style='font-weight:bold;color:#00b050;margin-bottom:8px;'>✅ Bu Hafta Tamamlanan</div>
        <ul style='margin:0;padding-left:18px;font-size:13px;'>{c_html}</ul>
      </div>
      <div style='flex:1;min-width:200px;background:#fff;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.08);padding:14px;border-left:5px solid #1976d2;'>
        <div style='font-weight:bold;color:#1976d2;margin-bottom:8px;'>🔵 Bu Hafta Marka Onayına Gelen</div>
        <ul style='margin:0;padding-left:18px;font-size:13px;'>{o_html}</ul>
      </div>
    </div>
    <div style='text-align:center;font-size:11px;color:#aaa;margin-top:20px;'>Haftalık rapor · bc-takip-production.up.railway.app</div>
  </div>
</body></html>"""


# ══════════════════════════════════════════════════════════════════════════
#  HATA UYARISI + MAİL
# ══════════════════════════════════════════════════════════════════════════

def send_error_alert(error: str, trigger: str = "?") -> None:
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"></head>
<body style='font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif;background:#f5f5f5;margin:0;padding:20px;'>
  <div style='max-width:560px;margin:0 auto;'>
    <div style='background:#c62828;color:#fff;border-radius:10px;padding:20px 24px;margin-bottom:16px;'>
      <div style='font-size:11px;text-transform:uppercase;letter-spacing:1px;opacity:.7;margin-bottom:4px;'>PunchBBDO — Excel Takip</div>
      <div style='font-size:22px;font-weight:bold;'>⚠️ Sistem Hatası</div>
      <div style='font-size:13px;opacity:.8;margin-top:4px;'>{now} · tetikleyen: {trigger}</div>
    </div>
    <div style='background:#fff;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.08);padding:16px;border-left:5px solid #c62828;'>
      <div style='font-weight:bold;color:#c62828;margin-bottom:8px;'>Hata Detayı</div>
      <pre style='background:#fff3f3;padding:12px;border-radius:4px;font-size:12px;white-space:pre-wrap;word-break:break-all;'>{error}</pre>
    </div>
  </div>
</body></html>"""
    try:
        send_email(f"⚠️ BC Takip Sistemi Hatası — {now}", f"Hata: {error}", html)
        print("✉️  Error alert gönderildi")
    except Exception as e:
        print(f"⚠️  Error alert gönderilemedi: {e}")


def send_email(subject: str, body_text: str, body_html: str) -> None:
    if not BREVO_API_KEY:
        raise ValueError("BREVO_API_KEY ayarlanmamış")
    payload = json.dumps({
        "sender":      {"name": "Excel Rapor", "email": RECIPIENT_EMAIL},
        "to":          [{"email": RECIPIENT_EMAIL}],
        "subject":     subject,
        "textContent": body_text,
        "htmlContent": body_html,
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.brevo.com/v3/smtp/email", data=payload, method="POST",
        headers={"api-key": BREVO_API_KEY, "Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            print(f"✉️  Brevo: {json.loads(r.read())}")
    except urllib.error.HTTPError as e:
        raise Exception(f"Brevo {e.code}: {e.read().decode('utf-8', errors='replace')}")


# ══════════════════════════════════════════════════════════════════════════
#  ANA RAPOR AKIŞI
# ══════════════════════════════════════════════════════════════════════════

def run_report(trigger: str = "cron") -> str:
    if not _report_lock.acquire(blocking=False):
        return "SKIPPED: rapor zaten çalışıyor"
    try:
        today = datetime.now().strftime("%d.%m.%Y %H:%M")
        print(f"\n▶  Rapor [{trigger}]: {today}\n{'─'*50}")
        try:
            return _run_report_inner(today, trigger)
        except Exception as e:
            err = str(e)
            print(f"❌ Kritik hata [{trigger}]: {err}")
            set_last_error(err)
            if BREVO_API_KEY:
                send_error_alert(err, trigger)
            return f"ERROR: {err}"
    finally:
        _report_lock.release()


def _run_report_inner(today: str, trigger: str) -> str:
    token = get_access_token()
    print("✅ Token alındı")

    all_todos = []
    for acct_id in BASECAMP_ACCOUNT_IDS:
        try:
            raw = bc_get(token, acct_id, "my/assignments.json")
            fetched = []
            for item in raw:
                if isinstance(item, dict) and "priorities" in item:
                    fetched.extend(item.get("priorities", []))
                    fetched.extend(item.get("non_priorities", []))
                elif isinstance(item, dict) and item.get("title"):
                    fetched.append(item)
            for t in fetched:
                t["_account_id"] = acct_id
            all_todos.extend(fetched)
            print(f"📋 Hesap {acct_id}: {len(fetched)} görev")
        except Exception as e:
            print(f"⚠️  Hesap {acct_id}: {e}")

    todos = [
        t for t in all_todos
        if (t.get("bucket") or {}).get("name", "").lower().strip() in TARGET_PROJECTS
        and not t.get("completed", False)
    ]
    print(f"🎯 Hedef projelerde aktif: {len(todos)}")

    excel_error = ""
    excel_tasks = []
    try:
        excel_tasks = read_excel_tasks()
        print(f"📊 Excel: {len(excel_tasks)} iş")
    except Exception as e:
        excel_error = str(e)
        print(f"⚠️  Excel: {e}")

    excel_by_id    = {t["todo_id"]: t for t in excel_tasks if t.get("todo_id")}
    excel_by_name  = {t["name"].lower(): t for t in excel_tasks}
    processed_keys = set()

    yesile_boya  = []
    renksiz_yap  = []
    sil_listesi  = []
    ekle_listesi = []
    url_eksik    = []

    # ── PASS 1: Ertuğ'a atanmış Basecamp todoları ───────────────────────
    for todo in todos:
        if is_produksiyon(todo) or is_sm_pm(todo):
            continue
        tid         = todo.get("id")
        name        = get_todo_title(todo)
        project_raw = (todo.get("bucket") or {}).get("name", "")
        brand       = TARGET_PROJECTS.get(project_raw.lower().strip(), project_raw)
        bc_onay     = is_marka_onayinda(todo)
        bucket_id_v = (todo.get("bucket") or {}).get("id")
        acct_id_v   = todo.get("_account_id", "")
        excel_item  = excel_by_id.get(tid) or excel_by_name.get(name.lower())

        if excel_item:
            key = excel_item.get("todo_id") or excel_item["name"].lower()
            processed_keys.add(key)
            is_green = excel_item.get("cell_color") == "green"
            if bc_onay and not is_green:
                yesile_boya.append({"name": name, "brand": brand,
                                    "todo_id": tid, "bucket_id": bucket_id_v, "url_account_id": acct_id_v})
                print(f"  🟢 [YEŞİLE BOYA] {name}")
            elif not bc_onay and is_green:
                renksiz_yap.append({"name": name, "brand": brand,
                                    "todo_id": tid, "bucket_id": bucket_id_v, "url_account_id": acct_id_v})
                print(f"  ⬜ [RENKSİZ YAP] {name}")
            else:
                print(f"  ✅ [DOĞRU RENK] {name}")
        else:
            ekle_listesi.append({"name": name, "brand": brand, "yesile_boya": bc_onay,
                                 "todo_id": tid, "bucket_id": bucket_id_v, "url_account_id": acct_id_v})
            print(f"  ➕ [EKLE] {name}")

    # ── PASS 2: Excel'de olup Ertuğ'a atanmamış todoları ────────────────
    for t in excel_tasks:
        key = t.get("todo_id") or t["name"].lower()
        if key in processed_keys:
            continue
        tid       = t.get("todo_id")
        bucket_id = t.get("bucket_id")
        url_acct  = t.get("url_account_id")
        is_green  = t.get("cell_color") == "green"

        if tid and bucket_id and url_acct:
            info = get_todo_info(token, url_acct, bucket_id, tid)
            if info is None:
                sil_listesi.append(t); print(f"  🗑️  [SİL - bulunamadı] {t['name']}")
            elif info["completed"]:
                sil_listesi.append(t); print(f"  🗑️  [SİL - tamamlandı] {t['name']}")
            elif info["produksiyon"]:
                print(f"  ⏭️  [SKIP - prodüksiyon] {t['name']}")
            elif any(k in info["list_name"] for k in SM_PM_LIST_KEYWORDS):
                print(f"  ⏭️  [SKIP - sm&pm] {t['name']}")
            elif "marka onay" in info["list_name"]:
                if not is_green:
                    yesile_boya.append({**t}); print(f"  🟢 [YEŞİLE BOYA - başka kişi] {t['name']}")
                else:
                    print(f"  ✅ [DOĞRU RENK - başka kişi] {t['name']}")
            else:
                if is_green:
                    renksiz_yap.append({**t}); print(f"  ⬜ [RENKSİZ YAP - başka kişi] {t['name']}")
                else:
                    print(f"  ✅ [DOĞRU RENK - başka kişi] {t['name']}")
        else:
            url_eksik.append(t); print(f"  🔗 [URL EKSİK] {t['name']}")

    prev_state = load_state()
    changes    = compute_changes(prev_state, sil_listesi, yesile_boya, renksiz_yap, ekle_listesi)
    first_seen = save_state(sil_listesi, yesile_boya, renksiz_yap, ekle_listesi, url_eksik, today)

    # Değişiklikleri kalıcı changelog dosyasına yaz
    append_changelog(changes, today, trigger)

    report      = build_report(yesile_boya, renksiz_yap, sil_listesi, ekle_listesi, url_eksik, today, excel_error, changes, first_seen)
    report_html = build_html_report(yesile_boya, renksiz_yap, sil_listesi, ekle_listesi, url_eksik, today, excel_error, changes, first_seen)
    print(f"\n{'═'*50}\n{report}\n{'═'*50}")

    if BREVO_API_KEY:
        try:
            send_email(f"📋 Excel Güncelleme Talimatları — {today}", report, report_html)
            print("✉️  Mail gönderildi!")
        except Exception as e:
            print(f"⚠️  Mail: {e}")
            report += f"\n\n⚠️ Mail gönderilemedi: {e}"
    return report


def run_weekly_summary():
    print("\n📅 Haftalık özet başlatıldı")
    state = load_state()
    if not state:
        return
    today = datetime.now().strftime("%d.%m.%Y")
    if BREVO_API_KEY:
        try:
            send_email(f"📅 Haftalık Özet — {today}", f"Haftalık özet: {today}", build_weekly_html(state))
            print("✉️  Haftalık özet gönderildi")
        except Exception as e:
            print(f"⚠️  Haftalık özet: {e}")


# ══════════════════════════════════════════════════════════════════════════
#  SABAH DİGEST (09:00 — API çağrısı yok, state'den okur)
# ══════════════════════════════════════════════════════════════════════════

def build_digest_html(state: dict) -> str:
    today      = datetime.now().strftime("%d.%m.%Y")
    fs         = state.get("first_seen", {})
    ts         = state.get("timestamp", "—")

    sil_items     = _as_items(state.get("sil", []))
    yesile_items  = _as_items(state.get("yesile", []))
    renksiz_items = _as_items(state.get("renksiz", []))
    ekle_items    = _as_items(state.get("ekle", []))
    total         = len(sil_items) + len(yesile_items) + len(renksiz_items) + len(ekle_items)

    # Uzun bekleyenler
    long_waiters = []
    for key, date_str in fs.items():
        if not key.startswith("yesile:"):
            continue
        name = key[len("yesile:"):]
        try:
            days = (datetime.now().date() - datetime.strptime(date_str, "%Y-%m-%d").date()).days
            if days >= 3:
                long_waiters.append((name, days))
        except Exception:
            pass
    long_waiters.sort(key=lambda x: -x[1])

    def mini_card(label, count, color):
        return (f"<div style='flex:1;min-width:100px;background:#fff;border-radius:8px;"
                f"padding:12px 16px;text-align:center;box-shadow:0 1px 4px rgba(0,0,0,.08);"
                f"border-top:3px solid {color};'>"
                f"<div style='font-size:28px;font-weight:bold;color:{color};'>{count}</div>"
                f"<div style='font-size:11px;color:#888;margin-top:2px;'>{label}</div></div>")

    waiter_rows = "".join(
        f"<li style='padding:3px 0;'><b>{n}</b> — "
        f"<span style='color:{'#c62828' if d>=7 else '#f57c00'};font-weight:bold;'>{d} gün</span></li>"
        for n, d in long_waiters[:5]
    )
    waiter_block = (
        f"<div style='background:#fff8e1;border:1px solid #ffe082;border-radius:8px;"
        f"padding:12px 16px;margin-top:16px;'>"
        f"<div style='font-weight:bold;color:#f57c00;margin-bottom:6px;'>⏰ Uzun Bekleyenler ({len(long_waiters)} iş)</div>"
        f"<ul style='margin:0;padding-left:18px;font-size:13px;color:#555;'>{waiter_rows}</ul>"
        + (f"<div style='font-size:11px;color:#aaa;margin-top:4px;'>...ve {len(long_waiters)-5} daha</div>"
           if len(long_waiters) > 5 else "")
        + f"</div>"
    ) if long_waiters else ""

    hero_color = "#e53935" if total >= 5 else ("#f57c00" if total >= 2 else "#00b050")

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style='font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif;background:#f5f5f5;margin:0;padding:20px;'>
  <div style='max-width:560px;margin:0 auto;'>
    <div style='background:linear-gradient(135deg,#1a1a2e,#16213e);color:#fff;border-radius:10px;padding:20px 24px;margin-bottom:16px;'>
      <div style='font-size:11px;text-transform:uppercase;letter-spacing:1px;opacity:.7;margin-bottom:4px;'>PunchBBDO — Excel Takip</div>
      <div style='font-size:22px;font-weight:bold;'>☀️ Günlük Özet</div>
      <div style='font-size:13px;opacity:.8;margin-top:4px;'>{today} · Son rapor: {ts}</div>
    </div>
    <div style='background:#fff;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.08);padding:16px;margin-bottom:16px;border-left:5px solid {hero_color};'>
      <div style='font-size:13px;color:#666;margin-bottom:10px;'>Bekleyen aksiyonlar:</div>
      <div style='display:flex;gap:8px;flex-wrap:wrap;'>
        {mini_card("SİL", len(sil_items), "#e53935")}
        {mini_card("YEŞİLE BOYA", len(yesile_items), "#00b050")}
        {mini_card("RENKSİZ YAP", len(renksiz_items), "#757575")}
        {mini_card("EKLE", len(ekle_items), "#1976d2")}
      </div>
      {waiter_block}
    </div>
    <div style='text-align:center;'>
      <a href='https://bc-takip-production.up.railway.app/dashboard'
         style='background:#1976d2;color:#fff;padding:10px 24px;border-radius:6px;font-size:13px;font-weight:bold;text-decoration:none;'>
        📋 Dashboard'a Git
      </a>
    </div>
    <div style='text-align:center;font-size:11px;color:#aaa;margin-top:16px;'>Sabah özeti · bc-takip-production.up.railway.app</div>
  </div>
</body></html>"""


def run_morning_digest():
    """Hafta içi 09:00 — API çağrısı yapmadan mevcut state'i özetler."""
    print("\n☀️  Sabah digest başlatıldı")
    state = load_state()
    if not state:
        print("⚠️  State boş, digest atlandı")
        return
    today = datetime.now().strftime("%d.%m.%Y")
    if BREVO_API_KEY:
        try:
            sil   = len(_as_items(state.get("sil", [])))
            yesil = len(_as_items(state.get("yesile", [])))
            renks = len(_as_items(state.get("renksiz", [])))
            ekle  = len(_as_items(state.get("ekle", [])))
            total = sil + yesil + renks + ekle
            subject = f"☀️ Sabah Özeti — {today} ({total} aksiyon bekliyor)"
            send_email(subject, f"Bekleyen: {total} aksiyon", build_digest_html(state))
            print("✉️  Sabah digest gönderildi")
        except Exception as e:
            print(f"⚠️  Sabah digest: {e}")


# ══════════════════════════════════════════════════════════════════════════
#  DEADLINE TAKİBİ (09:01 — BC API'si gerektirir)
# ══════════════════════════════════════════════════════════════════════════

def build_deadline_html(due_today: list, due_tomorrow: list, fetched_at: str) -> str:
    def rows(items):
        if not items:
            return "<li style='color:#888;font-style:italic;'>Yok</li>"
        return "".join(
            f"<li style='padding:4px 0;border-bottom:1px solid #f0f0f0;'>"
            f"<b>{t['name']}</b>"
            f"<span style='color:#888;font-size:12px;'> — {t.get('brand','')}</span>"
            + (f" &nbsp;<a href='{t['url']}' style='font-size:11px;color:#1976d2;'>[BC ↗]</a>"
               if t.get("url") else "")
            + f"</li>"
            for t in items
        )

    today_block = (
        f"<div style='margin:12px 0;border-radius:8px;overflow:hidden;"
        f"box-shadow:0 1px 4px rgba(0,0,0,.08);border-left:5px solid #e53935;background:#fff;'>"
        f"<div style='background:#e53935;color:#fff;padding:10px 16px;font-weight:bold;font-size:14px;'>"
        f"🔴 Bugün Bitmesi Gereken ({len(due_today)})</div>"
        f"<ul style='margin:0;padding:12px 16px 12px 32px;list-style:disc;'>{rows(due_today)}</ul></div>"
    )
    tomorrow_block = (
        f"<div style='margin:12px 0;border-radius:8px;overflow:hidden;"
        f"box-shadow:0 1px 4px rgba(0,0,0,.08);border-left:5px solid #f57c00;background:#fff;'>"
        f"<div style='background:#f57c00;color:#fff;padding:10px 16px;font-weight:bold;font-size:14px;'>"
        f"🟡 Yarın Bitiyor ({len(due_tomorrow)})</div>"
        f"<ul style='margin:0;padding:12px 16px 12px 32px;list-style:disc;'>{rows(due_tomorrow)}</ul></div>"
    )
    total = len(due_today) + len(due_tomorrow)
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style='font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif;background:#f5f5f5;margin:0;padding:20px;'>
  <div style='max-width:600px;margin:0 auto;'>
    <div style='background:linear-gradient(135deg,#b71c1c,#c62828);color:#fff;border-radius:10px;padding:20px 24px;margin-bottom:16px;'>
      <div style='font-size:11px;text-transform:uppercase;letter-spacing:1px;opacity:.7;margin-bottom:4px;'>PunchBBDO — Excel Takip</div>
      <div style='font-size:22px;font-weight:bold;'>⏰ Deadline Uyarısı</div>
      <div style='font-size:13px;opacity:.8;margin-top:4px;'>{fetched_at} · {total} iş yaklaşıyor</div>
    </div>
    {today_block}
    {tomorrow_block}
    <div style='text-align:center;font-size:11px;color:#aaa;margin-top:20px;'>Deadline takibi · bc-takip-production.up.railway.app</div>
  </div>
</body></html>"""


def run_deadline_check(trigger: str = "cron") -> str:
    """Bugün ve yarın due_on olan Basecamp görevlerini bulup mail atar."""
    print(f"\n⏰ Deadline kontrolü [{trigger}]")
    try:
        token = get_access_token()
    except Exception as e:
        print(f"⚠️  Token: {e}")
        return f"ERROR: {e}"

    today_dt    = datetime.now().date()
    tomorrow_dt = today_dt + timedelta(days=1)
    due_today    = []
    due_tomorrow = []

    for acct_id in BASECAMP_ACCOUNT_IDS:
        try:
            raw = bc_get(token, acct_id, "my/assignments.json")
            todos = []
            for item in raw:
                if isinstance(item, dict) and "priorities" in item:
                    todos.extend(item.get("priorities", []))
                    todos.extend(item.get("non_priorities", []))
                elif isinstance(item, dict) and item.get("title"):
                    todos.append(item)
            for t in todos:
                if t.get("completed"):
                    continue
                proj_raw = (t.get("bucket") or {}).get("name", "")
                brand    = TARGET_PROJECTS.get(proj_raw.lower().strip(), "")
                if not brand:
                    continue  # sadece hedef projeler
                due_on = t.get("due_on")  # "2025-12-31" veya None
                if not due_on:
                    continue
                try:
                    due_date = datetime.strptime(due_on, "%Y-%m-%d").date()
                except Exception:
                    continue
                tid = t.get("id")
                bid = (t.get("bucket") or {}).get("id")
                url = (f"https://3.basecamp.com/{acct_id}/buckets/{bid}/todos/{tid}"
                       if tid and bid else "")
                entry = {"name": get_todo_title(t), "brand": brand,
                         "due_on": due_on, "url": url}
                if due_date == today_dt:
                    due_today.append(entry)
                    print(f"  🔴 [BUGÜN] {entry['name']}")
                elif due_date == tomorrow_dt:
                    due_tomorrow.append(entry)
                    print(f"  🟡 [YARIN] {entry['name']}")
        except Exception as e:
            print(f"⚠️  Deadline {acct_id}: {e}")

    total = len(due_today) + len(due_tomorrow)
    if total == 0:
        print("✅ Yaklaşan deadline yok")
        return "OK: deadline yok"

    fetched_at = datetime.now().strftime("%d.%m.%Y %H:%M")
    html = build_deadline_html(due_today, due_tomorrow, fetched_at)
    subject = (f"⏰ Deadline Uyarısı — {fetched_at} "
               f"({len(due_today)} bugün, {len(due_tomorrow)} yarın)")
    if BREVO_API_KEY:
        try:
            send_email(subject, f"Deadline: {len(due_today)} bugün, {len(due_tomorrow)} yarın", html)
            print("✉️  Deadline maili gönderildi")
        except Exception as e:
            print(f"⚠️  Deadline mail: {e}")
    return f"OK: {len(due_today)} bugün, {len(due_tomorrow)} yarın"


# ══════════════════════════════════════════════════════════════════════════
#  DEBOUNCE
# ══════════════════════════════════════════════════════════════════════════

def schedule_debounced_report(kind: str):
    global _debounce_timer
    with _debounce_lock:
        if _debounce_timer is not None and _debounce_timer.is_alive():
            _debounce_timer.cancel()
            print(f"⏱️  Debounce sıfırlandı [{kind}]")
        _debounce_timer = threading.Timer(DEBOUNCE_SECONDS, run_report, kwargs={"trigger": f"webhook:{kind}"})
        _debounce_timer.daemon = True
        _debounce_timer.start()
        print(f"⏱️  Debounce: {DEBOUNCE_SECONDS // 60} dk sonra [{kind}]")


# ══════════════════════════════════════════════════════════════════════════
#  FLASK ENDPOINT'LERİ
# ══════════════════════════════════════════════════════════════════════════

@app.route("/health")
def health():
    state = load_state()
    return jsonify({
        "status": "ok", "time": datetime.now().isoformat(),
        "last_report": state.get("timestamp", "—"),
        "last_error": state.get("last_error"),
        "state_dir": STATE_DIR,
        "webhook_secret": "set" if WEBHOOK_SECRET else "not_set",
    })


@app.route("/run")
def manual_run():
    return f"<pre>{run_report(trigger='manual')}</pre>", 200


WEBHOOK_TRIGGER_KINDS = {
    "todo_completed", "todo_uncompleted", "todo_created",
    "todo_assignment_changed", "todo_trashed", "todo_moved",
}


@app.route("/webhook", methods=["POST"])
def basecamp_webhook():
    if WEBHOOK_SECRET:
        if request.args.get("token", "") != WEBHOOK_SECRET:
            print("🔒 Webhook reddedildi: geçersiz token")
            return jsonify({"status": "unauthorized"}), 401

    payload = request.get_json(silent=True) or {}
    kind    = payload.get("kind", "unknown")

    if kind not in WEBHOOK_TRIGGER_KINDS:
        print(f"⏭️  Webhook atlandı [{kind}]")
        log_webhook_event(kind, "ignored")
        return jsonify({"status": "ignored", "kind": kind}), 200

    log_webhook_event(kind, "scheduled")
    schedule_debounced_report(kind)
    return jsonify({"status": "scheduled", "kind": kind, "delay_minutes": DEBOUNCE_SECONDS // 60}), 202


# ── /dashboard ─────────────────────────────────────────────────────────────

@app.route("/dashboard")
def dashboard():
    state = load_state()
    if not state:
        return ("<html><body style='font-family:sans-serif;padding:40px;'>"
                "<h2>Henüz rapor çalışmadı.</h2><a href='/run'>▶ Şimdi çalıştır</a>"
                "</body></html>"), 200

    ts          = state.get("timestamp", "—")
    last_error  = state.get("last_error")
    first_seen  = state.get("first_seen", {})
    history     = state.get("history", [])
    webhook_log = load_webhook_log()

    brand_filter = request.args.get("brand", "").lower().strip()

    def mbrand(item):
        return not brand_filter or item.get("brand", "").lower() == brand_filter

    sil_items      = [i for i in _as_items(state.get("sil", []))       if mbrand(i)]
    yesile_items   = [i for i in _as_items(state.get("yesile", []))    if mbrand(i)]
    renksiz_items  = [i for i in _as_items(state.get("renksiz", []))   if mbrand(i)]
    ekle_items     = [i for i in _as_items(state.get("ekle", []))      if mbrand(i)]
    url_eksik_items= [i for i in _as_items(state.get("url_eksik", [])) if mbrand(i)]

    total_action = len(sil_items) + len(yesile_items) + len(renksiz_items) + len(ekle_items)

    # Uzun bekleyenler
    brand_lkp = {i["name"]: i.get("brand", "") for i in _as_items(state.get("yesile", []))}
    long_waiters = []
    for key, date_str in first_seen.items():
        if not key.startswith("yesile:"):
            continue
        name = key[len("yesile:"):]
        ib = brand_lkp.get(name, "")
        if brand_filter and ib.lower() != brand_filter:
            continue
        try:
            days = (datetime.now().date() - datetime.strptime(date_str, "%Y-%m-%d").date()).days
            if days >= 3:
                long_waiters.append((name, ib, days))
        except Exception:
            pass
    long_waiters.sort(key=lambda x: -x[2])

    # ─── Hero metrics ────────────────────────────────────────────────────
    countdown   = next_report_countdown()
    avg_onay    = avg_onay_days(first_seen)
    hero_color  = "#e53935" if total_action >= 5 else ("#f57c00" if total_action >= 2 else "#00b050")

    # ─── Brand comparison ────────────────────────────────────────────────
    def brand_count(cat, brand):
        return sum(1 for i in _as_items(state.get(cat, []))
                   if i.get("brand", "").lower() == brand.lower())

    brands_table = (
        f"<div style='background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.1);"
        f"margin-bottom:20px;overflow:hidden;'>"
        f"<div style='background:#1a1a2e;color:#fff;padding:12px 16px;font-weight:bold;'>🏷️ Marka Karşılaştırması</div>"
        f"<table style='width:100%;border-collapse:collapse;'>"
        f"<tr style='background:#fafafa;'>"
        f"<th style='padding:8px 14px;text-align:left;font-size:12px;color:#888;'></th>"
        f"<th style='padding:8px 14px;text-align:center;font-size:12px;color:#888;'>🛒 Hopi</th>"
        f"<th style='padding:8px 14px;text-align:center;font-size:12px;color:#888;'>🚇 Metro</th>"
        f"</tr>"
        + "".join([
            f"<tr>"
            f"<td style='padding:8px 14px;border-top:1px solid #eee;font-size:13px;font-weight:bold;color:{c};'>{icon} {label}</td>"
            f"<td style='padding:8px 14px;border-top:1px solid #eee;text-align:center;font-size:16px;font-weight:bold;color:{c};'>{brand_count(cat, 'Hopi')}</td>"
            f"<td style='padding:8px 14px;border-top:1px solid #eee;text-align:center;font-size:16px;font-weight:bold;color:{c};'>{brand_count(cat, 'Metro')}</td>"
            f"</tr>"
            for cat, label, icon, c in [
                ("sil",     "SİL",         "🗑️", "#e53935"),
                ("yesile",  "YEŞİLE BOYA", "🟢", "#00b050"),
                ("renksiz", "RENKSİZ YAP", "⬜", "#757575"),
                ("ekle",    "EKLE",        "➕", "#1976d2"),
            ]
        ])
        + f"</table></div>"
    )

    # ─── Trend grafiği (CSS-only) ─────────────────────────────────────────
    last14 = sorted(history, key=lambda x: x.get("date", ""))[-14:]
    max_val = max((h.get("yesile_count", 0) + len(h.get("completed", [])) for h in last14), default=1)
    max_val = max(max_val, 1)

    bars_html = ""
    for h in last14:
        onay_h   = h.get("yesile_count", 0)
        compl_h  = len(h.get("completed", []))
        date_lbl = h.get("date", "")[-5:]   # MM-DD
        op  = min(100, round(onay_h  / max_val * 100))
        cp  = min(100, round(compl_h / max_val * 100))
        bars_html += (
            f"<div style='flex:1;display:flex;flex-direction:column;align-items:center;gap:2px;'>"
            f"<div style='width:100%;display:flex;flex-direction:column;justify-content:flex-end;height:60px;gap:1px;'>"
            f"<div title='Marka Onayına Geldi: {onay_h}' style='width:100%;height:{op}%;background:#00b050;border-radius:2px 2px 0 0;min-height:{2 if onay_h else 0}px;'></div>"
            f"<div title='Tamamlandı: {compl_h}' style='width:100%;height:{cp}%;background:#1976d2;border-radius:2px 2px 0 0;min-height:{2 if compl_h else 0}px;'></div>"
            f"</div>"
            f"<div style='font-size:9px;color:#aaa;margin-top:2px;'>{date_lbl}</div>"
            f"</div>"
        )

    trend_section = (
        f"<div style='background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.1);"
        f"padding:16px;margin-bottom:20px;'>"
        f"<div style='font-weight:bold;font-size:14px;color:#333;margin-bottom:10px;'>📈 14 Günlük Trend</div>"
        f"<div style='display:flex;gap:2px;align-items:flex-end;'>{bars_html}</div>"
        f"<div style='display:flex;gap:16px;margin-top:8px;'>"
        f"<span style='font-size:11px;color:#888;'><span style='color:#00b050;font-weight:bold;'>■</span> Marka Onayına Geldi</span>"
        f"<span style='font-size:11px;color:#888;'><span style='color:#1976d2;font-weight:bold;'>■</span> Tamamlandı</span>"
        f"</div></div>"
    ) if last14 else ""

    # ─── Webhook log ─────────────────────────────────────────────────────
    wh_rows = ""
    for ev in webhook_log:
        status_color = "#00b050" if ev.get("status") == "scheduled" else "#9e9e9e"
        wh_rows += (
            f"<tr>"
            f"<td style='padding:6px 12px;border-bottom:1px solid #eee;font-size:12px;color:#888;'>{ev.get('time','')}</td>"
            f"<td style='padding:6px 12px;border-bottom:1px solid #eee;font-size:12px;font-family:monospace;'>{ev.get('kind','')}</td>"
            f"<td style='padding:6px 12px;border-bottom:1px solid #eee;'>"
            f"<span style='background:{status_color};color:#fff;padding:1px 8px;border-radius:3px;font-size:11px;'>{ev.get('status','')}</span>"
            f"</td></tr>"
        )
    wh_section = ""
    if wh_rows:
        wh_section = (
            f"<div style='background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.1);"
            f"margin-bottom:20px;overflow:hidden;'>"
            f"<div style='background:#1a1a2e;color:#fff;padding:12px 16px;font-weight:bold;'>⚡ Son Webhook Aktivitesi</div>"
            f"<table style='width:100%;border-collapse:collapse;'>"
            f"<tr style='background:#fafafa;'>"
            f"<th style='padding:6px 12px;text-align:left;font-size:11px;color:#888;'>Zaman</th>"
            f"<th style='padding:6px 12px;text-align:left;font-size:11px;color:#888;'>Event</th>"
            f"<th style='padding:6px 12px;text-align:left;font-size:11px;color:#888;'>Durum</th>"
            f"</tr>{wh_rows}</table></div>"
        )

    # ─── Stat kartları ────────────────────────────────────────────────────
    def sc(label, count, color):
        return (f"<div style='background:#fff;border-radius:10px;padding:16px;text-align:center;"
                f"box-shadow:0 1px 4px rgba(0,0,0,.1);border-top:4px solid {color};'>"
                f"<div style='font-size:34px;font-weight:bold;color:{color};'>{count}</div>"
                f"<div style='font-size:12px;color:#666;margin-top:4px;'>{label}</div></div>")

    stat_cards = (
        f"<div style='display:grid;grid-template-columns:repeat(auto-fit,minmax(100px,1fr));gap:12px;margin-bottom:20px;'>"
        f"{sc('SİL',len(sil_items),'#e53935')}{sc('YEŞİLE BOYA',len(yesile_items),'#00b050')}"
        f"{sc('RENKSİZ YAP',len(renksiz_items),'#757575')}{sc('EKLE',len(ekle_items),'#1976d2')}"
        f"{sc('URL EKSİK',len(url_eksik_items),'#9c27b0')}</div>"
    )

    # ─── Uzun bekleyenler ────────────────────────────────────────────────
    waiter_rows = "".join(
        f"<tr><td style='padding:8px 12px;border-bottom:1px solid #eee;'>{n}</td>"
        f"<td style='padding:8px 12px;border-bottom:1px solid #eee;'>{b}</td>"
        f"<td style='padding:8px 12px;border-bottom:1px solid #eee;text-align:center;"
        f"color:{'#c62828' if d>=7 else '#f57c00'};font-weight:bold;'>{d} gün</td></tr>"
        for n, b, d in long_waiters
    )
    waiter_section = (
        f"<div style='background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.1);"
        f"margin-bottom:20px;overflow:hidden;'>"
        f"<div style='background:#f57c00;color:#fff;padding:12px 16px;font-weight:bold;'>"
        f"⏰ Uzun Süredir Marka Onayında — {len(long_waiters)} iş</div>"
        f"<table style='width:100%;border-collapse:collapse;'>"
        f"<tr style='background:#fafafa;'>"
        f"<th style='padding:8px 12px;text-align:left;font-size:12px;color:#888;'>İş Adı</th>"
        f"<th style='padding:8px 12px;text-align:left;font-size:12px;color:#888;'>Marka</th>"
        f"<th style='padding:8px 12px;text-align:center;font-size:12px;color:#888;'>Bekleme</th>"
        f"</tr>{waiter_rows}</table></div>"
    ) if long_waiters else ""

    # ─── Mevcut durum önizlemesi ─────────────────────────────────────────
    notes = load_notes()

    def preview_list(items, category, color, icon, label):
        if not items:
            return ""
        rows = ""
        for it in items:
            days      = get_days_in_category(first_seen, category, it["name"])
            badge     = _dur_badge_html(days)
            bc_link   = _bc_url(it)
            safe_name = it["name"].replace("'", "&#39;")
            name_html = (f"<a href='{bc_link}' target='_blank' style='color:#333;"
                         f"text-decoration:none;font-weight:bold;'>{it['name']} ↗</a>"
                         if bc_link else f"<b>{it['name']}</b>")
            note_info = notes.get(it["name"])
            note_html = ""
            if note_info:
                note_text = note_info.get("text", "")
                note_time = note_info.get("time", "")
                note_html = (f"<div style='margin-top:2px;font-size:11px;color:#7b1fa2;"
                             f"background:#f3e5f5;border-radius:3px;padding:2px 6px;display:inline-block;'>"
                             f"📝 {note_text}"
                             f"<span style='color:#bbb;margin-left:6px;'>{note_time}</span></div>")
            rows += (
                f"<li class='search-item' data-name='{safe_name}' "
                f"style='padding:5px 0;border-bottom:1px solid #f5f5f5;'>"
                f"{name_html} "
                f"<span style='color:#aaa;font-size:11px;'>— {it.get('brand','')}</span>"
                f"{badge}{note_html}</li>"
            )
        return (f"<div style='flex:1;min-width:220px;background:#fff;border-radius:8px;"
                f"box-shadow:0 1px 4px rgba(0,0,0,.08);overflow:hidden;border-left:4px solid {color};'>"
                f"<div style='background:{color};color:#fff;padding:8px 12px;"
                f"font-size:13px;font-weight:bold;'>{icon} {label} ({len(items)})</div>"
                f"<ul style='margin:0;padding:10px 12px 10px 28px;list-style:disc;font-size:13px;'>{rows}</ul></div>")

    preview_items = "".join(filter(None, [
        preview_list(sil_items,       "sil",       "#e53935", "🗑️", "SİL"),
        preview_list(yesile_items,    "yesile",    "#00b050", "🟢", "YEŞİLE BOYA"),
        preview_list(renksiz_items,   "renksiz",   "#757575", "⬜", "RENKSİZ YAP"),
        preview_list(ekle_items,      "ekle",      "#1976d2", "➕", "EKLE"),
        preview_list(url_eksik_items, "url_eksik", "#9c27b0", "🔗", "URL EKSİK"),
    ]))
    preview_section = (
        f"<details style='margin-bottom:20px;' open>"
        f"<summary style='cursor:pointer;background:#fff;border-radius:10px;"
        f"box-shadow:0 1px 4px rgba(0,0,0,.1);padding:14px 18px;"
        f"font-weight:bold;font-size:15px;list-style:none;'>"
        f"📋 Mevcut Durum{' — ' + brand_filter.title() if brand_filter else ''} ▾</summary>"
        f"<div style='display:flex;flex-wrap:wrap;gap:12px;margin-top:12px;'>{preview_items}</div>"
        f"</details>"
    ) if preview_items else ""

    # ─── Geçmiş tablosu ──────────────────────────────────────────────────
    hist_rows = ""
    for h in sorted(history, key=lambda x: x.get("date", ""), reverse=True)[:14]:
        compl = len(h.get("completed", []))
        onay  = len(h.get("new_onay", []))
        hist_rows += (
            f"<tr>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #eee;font-size:13px;'>{h.get('date','')}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #eee;text-align:center;'>"
            f"<span style='background:#00b050;color:#fff;border-radius:4px;padding:2px 8px;font-size:12px;'>{h.get('yesile_count',0)}</span></td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #eee;text-align:center;'>"
            f"<span style='background:#e53935;color:#fff;border-radius:4px;padding:2px 8px;font-size:12px;'>{h.get('sil_count',0)}</span></td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #eee;text-align:center;'>"
            f"<span style='background:#1976d2;color:#fff;border-radius:4px;padding:2px 8px;font-size:12px;'>{h.get('ekle_count',0)}</span></td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #eee;text-align:center;"
            f"font-weight:bold;color:#00b050;'>{'+ ' + str(compl) if compl else '—'}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #eee;text-align:center;"
            f"font-weight:bold;color:#1976d2;'>{'+ ' + str(onay) if onay else '—'}</td>"
            f"</tr>"
        )
    if not hist_rows:
        hist_rows = "<tr><td colspan='6' style='padding:16px;text-align:center;color:#888;'>Henüz veri yok</td></tr>"

    # ─── Filter butonları ─────────────────────────────────────────────────
    def fbtn(lbl, val):
        active = (brand_filter == val) or (val == "" and not brand_filter)
        bg = "#1976d2" if active else "rgba(255,255,255,.15)"
        href = f"/dashboard{'?brand=' + val if val else ''}"
        return (f"<a href='{href}' style='background:{bg};color:#fff;padding:5px 12px;"
                f"border-radius:5px;font-size:12px;text-decoration:none;'>{lbl}</a>")

    # ─── Hata bloğu ───────────────────────────────────────────────────────
    error_block = (
        f"<div style='background:#fff3f3;border:1px solid #ffcdd2;border-radius:10px;"
        f"padding:14px 18px;margin-bottom:16px;border-left:5px solid #c62828;'>"
        f"<div style='font-weight:bold;color:#c62828;margin-bottom:4px;'>"
        f"⚠️ Son Hata — {last_error.get('time','')}</div>"
        f"<pre style='margin:0;font-size:12px;white-space:pre-wrap;word-break:break-all;'>"
        f"{last_error.get('message','')}</pre></div>"
    ) if last_error else ""

    # ─── Sprint badge ──────────────────────────────────────────────────────
    active_sprint = get_active_sprint()
    if active_sprint:
        sp_name = active_sprint.get("name", "Sprint")
        sp_end  = active_sprint.get("end", "")
        sprint_badge_html = (
            f"<a href='/sprint' style='background:#7b1fa2;color:#fff;padding:5px 14px;"
            f"border-radius:5px;font-size:12px;text-decoration:none;display:inline-block;margin-top:6px;'>"
            f"🏃 {sp_name} &nbsp;→&nbsp; {sp_end}</a>"
        )
    else:
        sprint_badge_html = (
            f"<a href='/sprint/new' style='background:rgba(123,31,162,.25);color:#ce93d8;"
            f"padding:5px 14px;border-radius:5px;font-size:12px;text-decoration:none;"
            f"display:inline-block;margin-top:6px;'>+ Sprint / Kampanya tanımla</a>"
        )

    # ─── Tahmin widget'ı ───────────────────────────────────────────────────
    forecast_section = build_forecast_widget(history)

    # ─── Isı haritası ──────────────────────────────────────────────────────
    heatmap_section = build_heatmap_html(webhook_log)

    # ─── Not formu ─────────────────────────────────────────────────────────
    all_item_names = sorted({
        it["name"]
        for cat in ["sil", "yesile", "renksiz", "ekle", "url_eksik"]
        for it in _as_items(state.get(cat, []))
    })
    note_options = "<option value=''>— İş seç —</option>"
    for n in all_item_names:
        safe_n = n.replace("'", "&#39;").replace('"', "&quot;")
        existing = notes.get(n, {}).get("text", "")
        sel = " selected" if existing else ""
        note_options += f"<option value='{safe_n}'{sel}>{n}</option>"

    notes_form_section = (
        f"<details style='margin-bottom:20px;'>"
        f"<summary style='cursor:pointer;background:#fff;border-radius:10px;"
        f"box-shadow:0 1px 4px rgba(0,0,0,.1);padding:14px 18px;"
        f"font-weight:bold;font-size:14px;list-style:none;color:#7b1fa2;'>📝 İş Notları ▾</summary>"
        f"<div style='background:#fff;border-radius:0 0 10px 10px;padding:16px 18px;"
        f"box-shadow:0 2px 4px rgba(0,0,0,.08);margin-top:-4px;'>"
        f"<form method='POST' action='/note' style='display:flex;flex-wrap:wrap;gap:10px;align-items:flex-end;'>"
        f"<div style='flex:2;min-width:180px;'>"
        f"<label style='font-size:12px;color:#888;display:block;margin-bottom:4px;'>İş Adı</label>"
        f"<select name='item_name' style='width:100%;padding:8px 10px;border:1px solid #ddd;"
        f"border-radius:6px;font-size:13px;'>{note_options}</select></div>"
        f"<div style='flex:3;min-width:200px;'>"
        f"<label style='font-size:12px;color:#888;display:block;margin-bottom:4px;'>Not (boş bırakırsan silinir)</label>"
        f"<input type='text' name='note_text' id='noteInput' placeholder='Örn: kliyente soruldu, revizyon bekleniyor...' maxlength='200'"
        f" style='width:100%;box-sizing:border-box;padding:8px 10px;border:1px solid #ddd;"
        f"border-radius:6px;font-size:13px;'></div>"
        f"<button type='submit' style='padding:8px 20px;background:#7b1fa2;color:#fff;"
        f"border:none;border-radius:6px;font-size:13px;cursor:pointer;font-weight:bold;'>Kaydet</button>"
        f"</form>"
        + _build_notes_list(notes)
        + f"</div></details>"
    )

    return f"""<!DOCTYPE html>
<html><head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <meta http-equiv="refresh" content="300">
  <title>BC Takip · Dashboard</title>
  <style>
    body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif;background:#f0f2f5;margin:0;padding:20px;transition:background .2s,color .2s;}}
    .wrap{{max-width:860px;margin:0 auto;}}
    a{{color:#1976d2;text-decoration:none;}} a:hover{{text-decoration:underline;}}
    details>summary::-webkit-details-marker{{display:none;}}
    /* ── Dark mode ── */
    body.dark{{filter:invert(100%) hue-rotate(180deg);}}
    body.dark img, body.dark [style*="background:linear-gradient"]{{filter:invert(100%) hue-rotate(180deg);}}
    /* ── Arama ── */
    #search-bar{{width:100%;box-sizing:border-box;padding:9px 14px;border:1px solid #ddd;border-radius:8px;
      font-size:14px;margin-bottom:14px;background:#fff;outline:none;}}
    #search-bar:focus{{border-color:#1976d2;box-shadow:0 0 0 3px rgba(25,118,210,.15);}}
    .search-item{{transition:opacity .15s;}}
    .search-item.hidden{{display:none;}}
  </style>
</head>
<body><div class="wrap">

  <div style='background:linear-gradient(135deg,#1a1a2e,#16213e);color:#fff;border-radius:12px;padding:22px 26px;margin-bottom:20px;'>
    <div style='display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:12px;'>
      <div>
        <div style='font-size:11px;text-transform:uppercase;letter-spacing:1px;opacity:.7;'>PunchBBDO — Excel Takip</div>
        <div style='font-size:24px;font-weight:bold;margin-top:4px;'>📋 Dashboard</div>
        <div style='font-size:13px;opacity:.7;margin-top:4px;'>Son güncelleme: {ts} &nbsp;·&nbsp; 5dk'da bir yenilenir</div>
        <div style='display:flex;gap:6px;margin-top:10px;flex-wrap:wrap;'>
          {fbtn('Tümü','')}{fbtn('Hopi','hopi')}{fbtn('Metro','metro')}
        </div>
        {sprint_badge_html}
      </div>
      <div style='display:flex;gap:8px;flex-wrap:wrap;align-items:flex-start;'>
        <a href='/run' style='background:#00b050;color:#fff;padding:8px 16px;border-radius:6px;font-size:13px;font-weight:bold;text-decoration:none;'>▶ Rapor Çalıştır</a>
        <a href='/status' style='background:rgba(255,255,255,.15);color:#fff;padding:8px 16px;border-radius:6px;font-size:13px;text-decoration:none;'>📡 BC Durumu</a>
        <a href='/history' style='background:rgba(255,255,255,.15);color:#fff;padding:8px 16px;border-radius:6px;font-size:13px;text-decoration:none;'>📅 Geçmiş</a>
        <a href='/export.csv' style='background:rgba(255,255,255,.15);color:#fff;padding:8px 16px;border-radius:6px;font-size:13px;text-decoration:none;'>⬇️ CSV</a>
        <button onclick="toggleDark()" id="darkBtn"
          style='background:rgba(255,255,255,.15);color:#fff;padding:8px 16px;border-radius:6px;font-size:13px;border:none;cursor:pointer;'>🌙 Dark</button>
      </div>
    </div>
  </div>

  <input id="search-bar" type="text" placeholder="🔍 İş ara... (anlık filtre)" oninput="doSearch(this.value)">

  <script>
  // Dark mode
  (function(){{
    if(localStorage.getItem('bc_dark')==='1'){{
      document.body.classList.add('dark');
      var b=document.getElementById('darkBtn');
      if(b)b.textContent='☀️ Light';
    }}
  }})();
  function toggleDark(){{
    var on=document.body.classList.toggle('dark');
    localStorage.setItem('bc_dark',on?'1':'0');
    var b=document.getElementById('darkBtn');
    if(b)b.textContent=on?'☀️ Light':'🌙 Dark';
  }}
  // Search
  function doSearch(q){{
    q=q.trim().toLowerCase();
    document.querySelectorAll('.search-item').forEach(function(el){{
      if(!q){{el.classList.remove('hidden');return;}}
      var name=(el.dataset.name||'').toLowerCase();
      el.classList.toggle('hidden',!name.includes(q));
    }});
  }}
  </script>

  {error_block}

  <div style='display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;margin-bottom:20px;'>
    <div style='background:#fff;border-radius:10px;padding:18px;box-shadow:0 1px 4px rgba(0,0,0,.1);border-top:4px solid {hero_color};text-align:center;'>
      <div style='font-size:42px;font-weight:bold;color:{hero_color};'>{total_action}</div>
      <div style='font-size:12px;color:#666;margin-top:4px;'>Toplam Aksiyon</div>
    </div>
    <div style='background:#fff;border-radius:10px;padding:18px;box-shadow:0 1px 4px rgba(0,0,0,.1);text-align:center;'>
      <div style='font-size:16px;font-weight:bold;color:#333;margin-top:6px;'>{countdown}</div>
      <div style='font-size:12px;color:#666;margin-top:4px;'>Sonraki Rapor</div>
    </div>
    <div style='background:#fff;border-radius:10px;padding:18px;box-shadow:0 1px 4px rgba(0,0,0,.1);text-align:center;'>
      <div style='font-size:24px;font-weight:bold;color:#f57c00;margin-top:4px;'>{avg_onay}</div>
      <div style='font-size:12px;color:#666;margin-top:4px;'>Ort. Onay Bekleme</div>
    </div>
  </div>

  {forecast_section}
  {stat_cards}
  {brands_table}
  {waiter_section}
  {trend_section}
  {heatmap_section}
  {preview_section}
  {notes_form_section}

  <div style='background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.1);overflow:hidden;margin-bottom:20px;'>
    <div style='background:#1a1a2e;color:#fff;padding:12px 16px;font-weight:bold;'>
      📅 Son 14 Gün Geçmiş
      <a href='/history' style='color:rgba(255,255,255,.6);font-size:12px;font-weight:normal;margin-left:8px;'>tümünü gör →</a>
    </div>
    <table style='width:100%;border-collapse:collapse;'>
      <tr style='background:#fafafa;'>
        <th style='padding:8px 12px;text-align:left;font-size:12px;color:#888;'>Tarih</th>
        <th style='padding:8px 12px;text-align:center;font-size:12px;color:#888;'>Onayda</th>
        <th style='padding:8px 12px;text-align:center;font-size:12px;color:#888;'>Silinecek</th>
        <th style='padding:8px 12px;text-align:center;font-size:12px;color:#888;'>Eklenecek</th>
        <th style='padding:8px 12px;text-align:center;font-size:12px;color:#888;'>Tamamlandı</th>
        <th style='padding:8px 12px;text-align:center;font-size:12px;color:#888;'>Yeni Onay</th>
      </tr>{hist_rows}
    </table>
  </div>

  {wh_section}

  <div style='text-align:center;font-size:11px;color:#aaa;margin-top:4px;'>
    bc-takip-production.up.railway.app ·
    <a href='/run'>Manuel</a> · <a href='/status'>BC Durumu</a> ·
    <a href='/history'>Geçmiş</a> · <a href='/sprint'>Sprint</a> ·
    <a href='/changelog'>Changelog</a> · <a href='/deadlines'>Deadline</a> ·
    <a href='/export.csv'>CSV</a> · <a href='/health'>Health</a> · <a href='/debug-excel'>Excel Debug</a>
  </div>
</div></body></html>""", 200


# ── /status ────────────────────────────────────────────────────────────────

@app.route("/status")
def status_page():
    """Salt Basecamp görünümü — Excel katmanı yok, my/assignments.json bazlı."""
    try:
        token = get_access_token()
    except Exception as e:
        return f"<pre>Token hatası: {e}</pre>", 500

    # Tüm aktif todoları çek, proje → liste → işler şeklinde grupla
    by_project = defaultdict(lambda: defaultdict(list))

    for acct_id in BASECAMP_ACCOUNT_IDS:
        try:
            raw = bc_get(token, acct_id, "my/assignments.json")
            for item in raw:
                todos = []
                if isinstance(item, dict) and "priorities" in item:
                    todos = item.get("priorities", []) + item.get("non_priorities", [])
                elif isinstance(item, dict) and item.get("title"):
                    todos = [item]
                for t in todos:
                    if t.get("completed"):
                        continue
                    proj = (t.get("bucket") or {}).get("name", "Diğer")
                    lst  = (t.get("parent") or {}).get("title", "—")
                    tid  = t.get("id")
                    bid  = (t.get("bucket") or {}).get("id")
                    by_project[proj][lst].append({
                        "name":    get_todo_title(t),
                        "todo_id": tid,
                        "bucket_id": bid,
                        "acct_id": acct_id,
                    })
        except Exception as e:
            print(f"⚠️  Status {acct_id}: {e}")

    def list_color(list_name: str) -> str:
        ln = list_name.lower()
        if "marka onay" in ln:
            return "#00b050"
        if any(k in ln for k in PRODUKSIYON_LIST_KEYWORDS):
            return "#f57c00"
        if any(k in ln for k in SM_PM_LIST_KEYWORDS):
            return "#9e9e9e"
        if "tasarım" in ln or "design" in ln or "hazırlık" in ln:
            return "#1976d2"
        return "#607d8b"

    total_active = sum(len(todos) for proj in by_project.values() for todos in proj.values())
    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")

    projects_html = ""
    for proj_name in sorted(by_project.keys()):
        brand_label = TARGET_PROJECTS.get(proj_name.lower().strip(), "")
        badge = (f"&nbsp;<span style='background:#1976d2;color:#fff;padding:1px 8px;"
                 f"border-radius:3px;font-size:11px;'>{brand_label}</span>"
                 if brand_label else "")

        lists_html = ""
        for list_name in sorted(by_project[proj_name].keys()):
            todos = by_project[proj_name][list_name]
            color = list_color(list_name)
            items_html = ""
            for t in todos:
                bc_link = (f"https://3.basecamp.com/{t['acct_id']}/buckets/{t['bucket_id']}/todos/{t['todo_id']}"
                           if t.get("todo_id") and t.get("bucket_id") else "")
                name_el = (f"<a href='{bc_link}' target='_blank' style='color:#333;text-decoration:none;'>"
                           f"{t['name']} <span style='color:#aaa;font-size:11px;'>↗</span></a>"
                           if bc_link else t["name"])
                items_html += f"<li style='padding:3px 0;border-bottom:1px solid #f5f5f5;'>{name_el}</li>"

            lists_html += (
                f"<div style='margin:0 0 10px 0;border-left:3px solid {color};padding-left:10px;'>"
                f"<div style='font-size:12px;font-weight:bold;color:{color};margin-bottom:4px;'>"
                f"{list_name} <span style='color:#aaa;font-weight:normal;'>({len(todos)})</span></div>"
                f"<ul style='margin:0;padding:0 0 0 16px;list-style:disc;font-size:13px;'>{items_html}</ul>"
                f"</div>"
            )

        proj_total = sum(len(t) for t in by_project[proj_name].values())
        projects_html += (
            f"<div style='background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.1);"
            f"margin-bottom:16px;overflow:hidden;'>"
            f"<div style='background:#1a1a2e;color:#fff;padding:12px 16px;font-weight:bold;display:flex;justify-content:space-between;align-items:center;'>"
            f"<span>📁 {proj_name}{badge}</span>"
            f"<span style='font-size:13px;opacity:.7;'>{proj_total} aktif iş</span>"
            f"</div>"
            f"<div style='padding:14px 16px;'>{lists_html}</div>"
            f"</div>"
        )

    if not projects_html:
        projects_html = "<div style='background:#fff;border-radius:10px;padding:24px;text-align:center;color:#888;'>Şu an aktif iş bulunamadı.</div>"

    return f"""<!DOCTYPE html>
<html><head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>BC Takip · Basecamp Durumu</title>
  <style>
    body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif;background:#f0f2f5;margin:0;padding:20px;}}
    .wrap{{max-width:760px;margin:0 auto;}}
    a{{color:#1976d2;text-decoration:none;}} a:hover{{text-decoration:underline;}}
  </style>
</head>
<body><div class="wrap">
  <div style='background:linear-gradient(135deg,#1a1a2e,#16213e);color:#fff;border-radius:12px;padding:22px 26px;margin-bottom:20px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px;'>
    <div>
      <div style='font-size:11px;text-transform:uppercase;letter-spacing:1px;opacity:.7;'>PunchBBDO — Excel Takip</div>
      <div style='font-size:24px;font-weight:bold;margin-top:4px;'>📡 Basecamp Durumu</div>
      <div style='font-size:13px;opacity:.7;margin-top:4px;'>{now_str} · {total_active} aktif iş · canlı veri</div>
    </div>
    <div style='display:flex;gap:8px;flex-wrap:wrap;'>
      <a href='/status' style='background:#00b050;color:#fff;padding:8px 14px;border-radius:6px;font-size:13px;font-weight:bold;text-decoration:none;'>🔄 Yenile</a>
      <a href='/dashboard' style='background:rgba(255,255,255,.15);color:#fff;padding:8px 14px;border-radius:6px;font-size:13px;text-decoration:none;'>← Dashboard</a>
    </div>
  </div>
  <div style='background:#e8f5e9;border-radius:8px;padding:10px 16px;margin-bottom:16px;font-size:13px;color:#2e7d32;border-left:4px solid #00b050;'>
    ℹ️ Bu sayfa sadece Basecamp'teki aktif işleri gösterir. Excel karşılaştırması, renk bilgisi veya aksiyon önerisi içermez.
  </div>
  {projects_html}
  <div style='text-align:center;font-size:11px;color:#aaa;margin-top:16px;'>
    <a href='/dashboard'>Dashboard</a> · <a href='/run'>Rapor Çalıştır</a> · <a href='/history'>Geçmiş</a>
  </div>
</div></body></html>""", 200


# ── /history ───────────────────────────────────────────────────────────────

@app.route("/history")
def history_page():
    state       = load_state()
    history     = state.get("history", [])
    date_filter = request.args.get("date", "").strip()

    if date_filter:
        entries    = [h for h in history if h.get("date") == date_filter]
        page_title = f"Geçmiş — {date_filter}"
    else:
        entries    = sorted(history, key=lambda x: x.get("date", ""), reverse=True)
        page_title = "Son 14 Gün Geçmiş"

    table_rows = ""
    for h in entries:
        compl    = h.get("completed", [])
        new_onay = h.get("new_onay", [])
        h_date = h.get("date", "")
        table_rows += (
            f"<tr>"
            f"<td style='padding:10px 14px;border-bottom:1px solid #eee;'>"
            f"<a href='/history?date={h_date}' style='font-weight:bold;color:#1976d2;'>{h.get('date','')}</a>"
            f"<br><span style='font-size:11px;color:#aaa;'>{h.get('time','')}</span></td>"
            f"<td style='padding:10px 14px;border-bottom:1px solid #eee;text-align:center;'>"
            f"<span style='background:#00b050;color:#fff;border-radius:4px;padding:2px 8px;font-size:13px;'>{h.get('yesile_count',0)}</span></td>"
            f"<td style='padding:10px 14px;border-bottom:1px solid #eee;text-align:center;'>"
            f"<span style='background:#e53935;color:#fff;border-radius:4px;padding:2px 8px;font-size:13px;'>{h.get('sil_count',0)}</span></td>"
            f"<td style='padding:10px 14px;border-bottom:1px solid #eee;text-align:center;'>"
            f"<span style='background:#1976d2;color:#fff;border-radius:4px;padding:2px 8px;font-size:13px;'>{h.get('ekle_count',0)}</span></td>"
            f"<td style='padding:10px 14px;border-bottom:1px solid #eee;font-size:13px;color:#333;'>{', '.join(compl) if compl else '—'}</td>"
            f"<td style='padding:10px 14px;border-bottom:1px solid #eee;font-size:13px;color:#333;'>{', '.join(new_onay) if new_onay else '—'}</td>"
            f"</tr>"
        )
    if not table_rows:
        table_rows = "<tr><td colspan='6' style='padding:24px;text-align:center;color:#888;'>Veri yok.</td></tr>"

    return f"""<!DOCTYPE html>
<html><head>
  <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>BC Takip · Geçmiş</title>
  <style>body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif;background:#f0f2f5;margin:0;padding:20px;}} .wrap{{max-width:860px;margin:0 auto;}} a{{color:#1976d2;text-decoration:none;}} a:hover{{text-decoration:underline;}}</style>
</head>
<body><div class="wrap">
  <div style='background:linear-gradient(135deg,#1a1a2e,#16213e);color:#fff;border-radius:12px;padding:22px 26px;margin-bottom:20px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;'>
    <div>
      <div style='font-size:11px;text-transform:uppercase;letter-spacing:1px;opacity:.7;'>PunchBBDO — Excel Takip</div>
      <div style='font-size:24px;font-weight:bold;margin-top:4px;'>📅 {page_title}</div>
    </div>
    <a href='/dashboard' style='background:rgba(255,255,255,.15);color:#fff;padding:8px 16px;border-radius:6px;font-size:13px;text-decoration:none;'>← Dashboard</a>
  </div>
  <div style='background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.1);overflow:hidden;'>
    <table style='width:100%;border-collapse:collapse;'>
      <tr style='background:#fafafa;'>
        <th style='padding:10px 14px;text-align:left;font-size:12px;color:#888;'>Tarih</th>
        <th style='padding:10px 14px;text-align:center;font-size:12px;color:#888;'>Onayda</th>
        <th style='padding:10px 14px;text-align:center;font-size:12px;color:#888;'>Silinecek</th>
        <th style='padding:10px 14px;text-align:center;font-size:12px;color:#888;'>Eklenecek</th>
        <th style='padding:10px 14px;text-align:left;font-size:12px;color:#888;'>Tamamlanan</th>
        <th style='padding:10px 14px;text-align:left;font-size:12px;color:#888;'>Yeni Onay</th>
      </tr>{table_rows}
    </table>
  </div>
  <div style='text-align:center;font-size:11px;color:#aaa;margin-top:20px;'>
    <a href='/dashboard'>Dashboard</a> · <a href='/run'>Manuel çalıştır</a> · <a href='/health'>Health</a>
  </div>
</div></body></html>""", 200


# ── Diğer endpoint'ler ─────────────────────────────────────────────────────

@app.route("/debug")
def debug():
    try:
        token = get_access_token()
    except Exception as e:
        return f"<pre>Token hatası: {e}</pre>", 500
    lines = []
    for acct_id in BASECAMP_ACCOUNT_IDS:
        try:
            raw = bc_get(token, acct_id, "my/assignments.json")
            todos = []
            for item in raw:
                if isinstance(item, dict) and "priorities" in item:
                    todos.extend(item.get("priorities", []))
                    todos.extend(item.get("non_priorities", []))
            lines.append(f"\n=== Hesap {acct_id} ({len(todos)} görev) ===")
            for t in todos:
                bucket = (t.get("bucket") or {}).get("name", "YOK")
                lst    = get_todolist_name(t)
                match  = "✅" if bucket.lower().strip() in TARGET_PROJECTS else "❌"
                lines.append(f"{match} [{bucket}] [{lst}] {get_todo_title(t)}")
        except Exception as e:
            lines.append(f"Hata: {e}")
    return f"<pre>{chr(10).join(lines)}</pre>", 200


# ── /api/state ─────────────────────────────────────────────────────────────

@app.route("/api/state")
def api_state():
    """Token korumalı JSON API — dışarıdan sorgulama için."""
    if WEBHOOK_SECRET:
        if request.args.get("token", "") != WEBHOOK_SECRET:
            return jsonify({"error": "unauthorized"}), 401
    state = load_state()
    if not state:
        return jsonify({"error": "no_data"}), 404
    # Sadece sayısal özet + listeler — ham state değil
    fs = state.get("first_seen", {})
    def enrich(items, cat):
        result = []
        for it in _as_items(items):
            d = get_days_in_category(fs, cat, it["name"])
            result.append({
                "name":       it["name"],
                "brand":      it.get("brand", ""),
                "days":       d,
                "bc_url":     _bc_url(it),
                "note":       load_notes().get(it["name"], {}).get("text", ""),
            })
        return result

    return jsonify({
        "timestamp":   state.get("timestamp", ""),
        "sil":         enrich(state.get("sil", []),       "sil"),
        "yesile":      enrich(state.get("yesile", []),    "yesile"),
        "renksiz":     enrich(state.get("renksiz", []),   "renksiz"),
        "ekle":        enrich(state.get("ekle", []),      "ekle"),
        "url_eksik":   enrich(state.get("url_eksik", []), "url_eksik"),
        "active_sprint": get_active_sprint(),
        "summary": {
            "total_action": (len(state.get("sil", [])) + len(state.get("yesile", [])) +
                             len(state.get("renksiz", [])) + len(state.get("ekle", []))),
            "avg_onay_days": avg_onay_days(fs),
        },
    })


# ── /note ──────────────────────────────────────────────────────────────────

@app.route("/note", methods=["POST", "GET"])
def note_endpoint():
    """Not ekle/güncelle (POST form) veya sil (GET ?delete=...)."""
    if request.method == "GET":
        item_name = request.args.get("delete", "").strip()
        if item_name:
            save_note(item_name, "")   # boş = sil
        return ("<script>history.back();</script>"
                "<a href='/dashboard'>← Dashboard</a>"), 200

    item_name = (request.form.get("item_name") or "").strip()
    note_text = (request.form.get("note_text") or "").strip()
    if item_name:
        save_note(item_name, note_text)
    return ("<script>history.back();</script>"
            "<a href='/dashboard'>← Dashboard</a>"), 200


# ── /sprint ────────────────────────────────────────────────────────────────

@app.route("/sprint")
def sprint_page():
    """Sprint özet sayfası. ?start=YYYY-MM-DD&end=YYYY-MM-DD ile belirli sprint."""
    state   = load_state()
    history = state.get("history", []) if state else []

    start_q = request.args.get("start", "").strip()
    end_q   = request.args.get("end", "").strip()

    if start_q and end_q:
        sprint = {"name": f"{start_q} → {end_q}", "start": start_q, "end": end_q}
    else:
        sprint = get_active_sprint()
        if not sprint:
            sprints = load_sprints()
            sprint  = sprints[-1] if sprints else None

    if not sprint:
        return (f"<html><body style='font-family:sans-serif;padding:40px;'>"
                f"<h2>Henüz sprint tanımlanmamış.</h2>"
                f"<a href='/sprint/new'>+ Yeni Sprint Oluştur</a> &nbsp;·&nbsp; "
                f"<a href='/dashboard'>← Dashboard</a></body></html>"), 200

    return build_sprint_page(sprint, history), 200


@app.route("/sprint/new", methods=["GET", "POST"])
def sprint_new():
    """Sprint oluşturma formu."""
    if request.method == "POST":
        name  = (request.form.get("name") or "").strip()
        start = (request.form.get("start") or "").strip()
        end   = (request.form.get("end") or "").strip()
        if name and start and end:
            save_sprint(name, start, end)
            return ("<script>window.location='/sprint';</script>"
                    "<a href='/sprint'>Sprint sayfasına git →</a>"), 200
        return "<p>Tüm alanları doldurun.</p><a href='/sprint/new'>Geri</a>", 400

    today     = datetime.now().strftime("%Y-%m-%d")
    two_weeks = (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d")
    return f"""<!DOCTYPE html>
<html><head>
  <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Yeni Sprint</title>
  <style>body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:#f0f2f5;margin:0;padding:40px;}}
  .card{{max-width:460px;margin:0 auto;background:#fff;border-radius:12px;padding:28px;box-shadow:0 2px 8px rgba(0,0,0,.1);}}
  label{{font-size:12px;color:#888;display:block;margin-bottom:4px;margin-top:14px;}}
  input{{width:100%;box-sizing:border-box;padding:9px 12px;border:1px solid #ddd;border-radius:6px;font-size:14px;}}
  input:focus{{outline:none;border-color:#7b1fa2;box-shadow:0 0 0 3px rgba(123,31,162,.15);}}
  button{{margin-top:20px;width:100%;padding:11px;background:#7b1fa2;color:#fff;border:none;border-radius:6px;font-size:14px;font-weight:bold;cursor:pointer;}}
  </style>
</head>
<body>
  <div class="card">
    <div style='font-size:22px;font-weight:bold;color:#7b1fa2;margin-bottom:4px;'>🏃 Yeni Sprint / Kampanya</div>
    <div style='font-size:13px;color:#888;margin-bottom:20px;'>
      History verisi bu tarih aralığına göre filtrelenecek.
    </div>
    <form method="POST" action="/sprint/new">
      <label>Sprint / Kampanya Adı</label>
      <input type="text" name="name" placeholder="Örn: Mayıs Kampanyası, Q2 Sprint" required>
      <label>Başlangıç Tarihi</label>
      <input type="date" name="start" value="{today}" required>
      <label>Bitiş Tarihi</label>
      <input type="date" name="end" value="{two_weeks}" required>
      <button type="submit">Sprint Oluştur</button>
    </form>
    <div style='text-align:center;margin-top:16px;font-size:12px;'>
      <a href='/sprint' style='color:#7b1fa2;'>← Mevcut Sprintler</a> &nbsp;·&nbsp;
      <a href='/dashboard' style='color:#1976d2;'>Dashboard</a>
    </div>
  </div>
</body></html>""", 200


# ── /changelog ─────────────────────────────────────────────────────────────

@app.route("/changelog")
def changelog_page():
    """Otomatik kaydedilen değişiklik geçmişi."""
    try:
        with open(CHANGELOG_FILE, encoding="utf-8") as f:
            raw = f.read()
    except FileNotFoundError:
        raw = "_Henüz değişiklik kaydedilmemiş. Bir rapor çalıştırıldıktan sonra burada görünür._"
    except Exception as e:
        raw = f"Dosya okunamadı: {e}"

    # Markdown-light render: ## → başlık, - → liste
    import html as _html
    lines   = raw.splitlines()
    out     = []
    in_list = False
    for line in lines:
        escaped = _html.escape(line)
        if line.startswith("## "):
            if in_list:
                out.append("</ul>"); in_list = False
            title = _html.escape(line[3:])
            out.append(f"<h3 style='margin:20px 0 6px;color:#1a1a2e;font-size:14px;border-bottom:1px solid #eee;padding-bottom:4px;'>{title}</h3>")
        elif line.startswith("- "):
            if not in_list:
                out.append("<ul style='margin:4px 0;padding-left:20px;font-size:13px;color:#444;'>")
                in_list = True
            out.append(f"<li style='padding:2px 0;'>{_html.escape(line[2:])}</li>")
        elif line.startswith("_") and line.endswith("_"):
            if in_list:
                out.append("</ul>"); in_list = False
            out.append(f"<p style='color:#888;font-style:italic;font-size:13px;'>{_html.escape(line[1:-1])}</p>")
        elif line.strip():
            if in_list:
                out.append("</ul>"); in_list = False
    if in_list:
        out.append("</ul>")

    content_html = "\n".join(out) or "<p style='color:#888;'>Değişiklik yok.</p>"
    entry_count  = raw.count("## ")

    return f"""<!DOCTYPE html>
<html><head>
  <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>BC Takip · Changelog</title>
  <style>body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:#f0f2f5;margin:0;padding:20px;}}
  .wrap{{max-width:760px;margin:0 auto;}} a{{color:#1976d2;text-decoration:none;}} a:hover{{text-decoration:underline;}}</style>
</head>
<body><div class="wrap">
  <div style='background:linear-gradient(135deg,#1a1a2e,#16213e);color:#fff;border-radius:12px;padding:22px 26px;margin-bottom:20px;
              display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px;'>
    <div>
      <div style='font-size:11px;text-transform:uppercase;letter-spacing:1px;opacity:.7;'>PunchBBDO — Excel Takip</div>
      <div style='font-size:24px;font-weight:bold;margin-top:4px;'>📜 Changelog</div>
      <div style='font-size:13px;opacity:.7;margin-top:4px;'>{entry_count} değişiklik kaydı</div>
    </div>
    <a href='/dashboard' style='background:rgba(255,255,255,.15);color:#fff;padding:8px 16px;border-radius:6px;font-size:13px;text-decoration:none;'>← Dashboard</a>
  </div>
  <div style='background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.1);padding:20px 24px;'>
    {content_html}
  </div>
  <div style='text-align:center;font-size:11px;color:#aaa;margin-top:16px;'>
    <a href='/dashboard'>Dashboard</a> · <a href='/history'>Geçmiş</a>
  </div>
</div></body></html>""", 200


@app.route("/export.csv")
def export_csv():
    """Mevcut state'i CSV olarak indir."""
    import csv
    import io as _io
    state = load_state()
    if not state:
        from flask import Response
        return Response("Henüz rapor çalışmadı.", mimetype="text/plain"), 404

    fs  = state.get("first_seen", {})
    buf = _io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Kategori", "İş Adı", "Marka", "Basecamp URL", "İlk Görülme"])
    for cat, label in [("sil","SİL"), ("yesile","YEŞİLE"), ("renksiz","RENKSİZ"),
                       ("ekle","EKLE"), ("url_eksik","URL_EKSİK")]:
        for item in _as_items(state.get(cat, [])):
            url   = _bc_url(item)
            first = fs.get(f"{cat}:{item['name']}", "")
            writer.writerow([label, item["name"], item.get("brand",""), url, first])

    filename = "bc-takip-" + datetime.now().strftime("%Y%m%d") + ".csv"
    from flask import Response
    return Response(
        "\ufeff" + buf.getvalue(),   # BOM → Excel Türkçe karakter desteği
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/deadlines")
def manual_deadlines():
    """Deadline kontrolünü manuel tetikler."""
    result = run_deadline_check(trigger="manual")
    return f"<pre>{result}</pre>", 200


@app.route("/debug-excel")
def debug_excel():
    try:
        tasks = read_excel_tasks()
    except Exception as e:
        return f"<pre>Excel hatası: {e}</pre>", 500
    lines = [f"Excel'den {len(tasks)} iş:\n"]
    for t in tasks:
        lines.append(f"[{t['brand']}] {t['name']}\n"
                     f"  todo_id={t.get('todo_id')} bucket_id={t.get('bucket_id')} "
                     f"acct={t.get('url_account_id')} color={t.get('cell_color')}\n")
    return f"<pre>{''.join(lines)}</pre>", 200


@app.route("/setup-webhooks")
def setup_webhooks():
    railway_url = (f"https://{request.host}/webhook?token={WEBHOOK_SECRET}"
                   if WEBHOOK_SECRET else f"https://{request.host}/webhook")
    results = []
    try:
        token = get_access_token()
    except Exception as e:
        return f"<pre>Token hatası: {e}</pre>", 500

    for acct_id in BASECAMP_ACCOUNT_IDS:
        try:
            projects = bc_get(token, acct_id, "projects.json")
        except Exception as e:
            results.append(f"❌ Hesap {acct_id}: {e}"); continue

        for proj in projects:
            if proj.get("name", "").lower().strip() not in TARGET_PROJECTS:
                continue
            proj_id = proj["id"]
            proj_name = proj["name"]
            try:
                existing = bc_get(token, acct_id, f"buckets/{proj_id}/webhooks.json")
                if any(w.get("payload_url") == railway_url for w in existing):
                    results.append(f"✅ {proj_name} — zaten kayıtlı"); continue
            except Exception:
                pass
            payload = json.dumps({"payload_url": railway_url}).encode()
            req = urllib.request.Request(
                f"https://3.basecampapi.com/{acct_id}/buckets/{proj_id}/webhooks.json",
                data=payload, method="POST",
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json",
                         "User-Agent": "IsOzetRaporu (ertugozerr@gmail.com)"},
            )
            try:
                with urllib.request.urlopen(req, timeout=30) as r:
                    resp = json.loads(r.read())
                    results.append(f"✅ {proj_name} — kaydedildi (ID: {resp.get('id')})")
            except Exception as e:
                results.append(f"❌ {proj_name}: {e}")

    note = f"\n🔒 WEBHOOK_SECRET aktif — URL'ye token eklendi." if WEBHOOK_SECRET else "\n⚠️  WEBHOOK_SECRET ayarlanmamış."
    return f"<pre>{chr(10).join(results)}\n\nWebhook URL: {railway_url}{note}</pre>", 200


# ══════════════════════════════════════════════════════════════════════════
#  SCHEDULER
# ══════════════════════════════════════════════════════════════════════════

def start_scheduler():
    scheduler = BackgroundScheduler(timezone="Europe/Istanbul")
    scheduler.add_job(
        run_report,
        CronTrigger(day_of_week="mon-fri", hour=18, minute=0, timezone="Europe/Istanbul"),
        kwargs={"trigger": "cron"}, id="daily_report", replace_existing=True,
    )
    scheduler.add_job(
        run_weekly_summary,
        CronTrigger(day_of_week="fri", hour=18, minute=5, timezone="Europe/Istanbul"),
        id="weekly_summary", replace_existing=True,
    )
    scheduler.add_job(
        run_morning_digest,
        CronTrigger(day_of_week="mon-fri", hour=9, minute=0, timezone="Europe/Istanbul"),
        id="morning_digest", replace_existing=True,
    )
    scheduler.add_job(
        run_deadline_check,
        CronTrigger(day_of_week="mon-fri", hour=9, minute=1, timezone="Europe/Istanbul"),
        kwargs={"trigger": "cron"}, id="deadline_check", replace_existing=True,
    )
    scheduler.start()
    print("📅 Scheduler:")
    print("   Haftaiçi 09:00 → Sabah digest (state özeti)")
    print("   Haftaiçi 09:01 → Deadline kontrolü (BC API)")
    print("   Haftaiçi 18:00 → Günlük rapor")
    print("   Cuma     18:05 → Haftalık özet")
    print(f"📁 State: {STATE_FILE}")
    print(f"🔒 Webhook: {'AKTIF' if WEBHOOK_SECRET else 'WEBHOOK_SECRET ayarlanmamış'}")


if __name__ == "__main__":
    start_scheduler()
    port = int(os.environ.get("PORT", 8080))
    print(f"🚀 Port {port}")
    app.run(host="0.0.0.0", port=port, debug=False)
